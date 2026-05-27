"""
PASO 6 (G8) — M&E Comparison Engine: Dafne vs Power BI.

Compares M&E calculated values between Dafne (Tbl_Integrado.xlsx) and
Power BI (Excel export or API output) to detect discrepancies and record
final decisions for output/outcome/impact indicators.

Decision history is persisted to:
  data/compare/{quarter}_decisions.json
"""
from __future__ import annotations

import json
import os
import pathlib
from datetime import datetime
from typing import Optional

import openpyxl

# ES: Directorio raíz del proyecto / EN: Project root directory
_PROJECT_ROOT = pathlib.Path(__file__).parent

# ES: Directorio de decisiones de comparación / EN: Comparison decisions directory
_COMPARE_DIR = _PROJECT_ROOT / "data" / "compare"

# ES: Tolerancia numérica para comparación / EN: Numeric tolerance for comparison
_TOLERANCE = 0.01

# ES: Hojas requeridas en Tbl_Integrado.xlsx / EN: Required sheets
_SHEET_ACTIVIDADES = "Tbl_Actvdd_m_e_data"
_SHEET_BENEFICIARIOS = "Tbl_Beneficiarios"
_SHEET_ORG = "Tbl_Org"


# ─────────────────────────────────────────────────────────────────────────────
# Data loaders
# ─────────────────────────────────────────────────────────────────────────────

def load_dafne_values(integrado_path: str) -> dict:
    """
    Extract output/outcome/impact values from Tbl_Integrado.xlsx using openpyxl.

    Returns:
        {
            "output": {
                "area_ha":         float  — sum of Area_ha from Tbl_Actvdd_m_e_data,
                "beneficiarios":   int    — row count from Tbl_Beneficiarios,
                "organizaciones":  int    — row count from Tbl_Org
            },
            "outcome": {
                "<CdgActvdd>": {"area_ha": float, "pct_logro": float}, ...
            },
            "impact": {
                "total_area_ha": float,
                "kpi_threshold": float  — None if not present
            },
            "rows": {
                "<CdgActvdd>": {"Area_ha": float, ...}  — raw row data from Tbl_Actvdd_m_e_data
            },
            "error": str | None
        }

    # ES: Extrae valores M&E de Tbl_Integrado.xlsx para comparación.
    # EN: Extracts M&E values from Tbl_Integrado.xlsx for comparison.
    """
    fp = pathlib.Path(integrado_path)
    if not fp.exists():
        return {"error": f"Archivo no encontrado / File not found: {integrado_path}"}

    try:
        wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"error": f"No se pudo abrir el archivo / Could not open file: {exc}"}

    result: dict = {"error": None, "rows": {}}

    try:
        # ── Output: Area_ha sum from Tbl_Actvdd_m_e_data ──────────────────
        area_ha_total = 0.0
        outcome_rows: dict = {}

        if _SHEET_ACTIVIDADES in wb.sheetnames:
            ws = wb[_SHEET_ACTIVIDADES]
            headers = [c.value for c in next(ws.iter_rows(max_row=1))]

            cdg_col = _col_idx(headers, ["CdgActvdd", "Cdg_Actvdd", "Codigo"])
            area_col = _col_idx(headers, ["Area_ha", "area_ha", "Area"])

            for row in ws.iter_rows(min_row=2, values_only=True):
                if cdg_col is None or area_col is None:
                    break
                cdg = row[cdg_col] if cdg_col < len(row) else None
                area = row[area_col] if area_col < len(row) else None
                if cdg is None:
                    continue
                area_val = float(area) if area is not None else 0.0
                area_ha_total += area_val

                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                outcome_rows[str(cdg)] = row_dict
                result["rows"][str(cdg)] = row_dict

        # ── Output: Beneficiarios count ────────────────────────────────────
        bnf_count = 0
        if _SHEET_BENEFICIARIOS in wb.sheetnames:
            ws = wb[_SHEET_BENEFICIARIOS]
            for _ in ws.iter_rows(min_row=2, values_only=True):
                bnf_count += 1

        # ── Output: Organizaciones count ───────────────────────────────────
        org_count = 0
        if _SHEET_ORG in wb.sheetnames:
            ws = wb[_SHEET_ORG]
            for _ in ws.iter_rows(min_row=2, values_only=True):
                org_count += 1

        result["output"] = {
            "area_ha": round(area_ha_total, 4),
            "beneficiarios": bnf_count,
            "organizaciones": org_count,
        }

        # ── Outcome: per-activity achievement ─────────────────────────────
        result["outcome"] = {
            cdg: {
                "area_ha": float(row_data.get("Area_ha") or row_data.get("area_ha") or 0.0),
                "pct_logro": float(row_data.get("Pct_Logro") or row_data.get("pct_logro") or 0.0),
            }
            for cdg, row_data in outcome_rows.items()
        }

        # ── Impact: total area and optional KPI threshold ──────────────────
        result["impact"] = {
            "total_area_ha": round(area_ha_total, 4),
            "kpi_threshold": None,
        }

    finally:
        wb.close()

    return result


def load_pbi_values(pbi_source: dict) -> dict:
    """
    Load PBI values from a confirmed Excel export or inline dict (G9 approach).

    pbi_source may be:
        {"type": "excel", "path": "...", "sheet": "..."}   — openpyxl read
        {"type": "inline", "output": {...}, "outcome": {...}, "impact": {...}}

    Returns same structure as load_dafne_values (without "rows").

    # ES: Carga valores PBI desde exportación Excel o dict en línea (enfoque G9).
    # EN: Loads PBI values from Excel export or inline dict (G9 approach).
    """
    src_type = pbi_source.get("type", "inline")

    if src_type == "inline":
        return {
            "output": pbi_source.get("output", {}),
            "outcome": pbi_source.get("outcome", {}),
            "impact": pbi_source.get("impact", {}),
            "rows": {},
            "error": None,
        }

    if src_type == "excel":
        path = pbi_source.get("path", "")
        sheet = pbi_source.get("sheet", "")
        fp = pathlib.Path(path)
        if not fp.exists():
            return {"error": f"Archivo PBI no encontrado / PBI file not found: {path}"}

        try:
            wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return {"error": f"No se pudo abrir PBI Excel / Could not open PBI Excel: {exc}"}

        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]

        area_ha_total = 0.0
        bnf_count = 0
        org_count = 0
        outcome: dict = {}
        rows: dict = {}

        cdg_col = _col_idx(headers, ["CdgActvdd", "Cdg_Actvdd", "Codigo"])
        area_col = _col_idx(headers, ["Area_ha", "area_ha", "Area"])

        for row in ws.iter_rows(min_row=2, values_only=True):
            if cdg_col is None or area_col is None:
                break
            cdg = row[cdg_col] if cdg_col < len(row) else None
            area = row[area_col] if area_col < len(row) else None
            if cdg is None:
                continue
            area_val = float(area) if area is not None else 0.0
            area_ha_total += area_val
            row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
            outcome[str(cdg)] = {"area_ha": area_val, "pct_logro": 0.0}
            rows[str(cdg)] = row_dict

        wb.close()

        return {
            "output": {
                "area_ha": round(area_ha_total, 4),
                "beneficiarios": bnf_count,
                "organizaciones": org_count,
            },
            "outcome": outcome,
            "impact": {"total_area_ha": round(area_ha_total, 4), "kpi_threshold": None},
            "rows": rows,
            "error": None,
        }

    return {"error": f"Tipo de fuente PBI desconocido / Unknown PBI source type: {src_type}"}


# ─────────────────────────────────────────────────────────────────────────────
# Comparison engine
# ─────────────────────────────────────────────────────────────────────────────

def compare_me_values(
    dafne_data: dict,
    pbi_data: dict,
    metrics: Optional[list] = None,
) -> dict:
    """
    Compare output/outcome/impact indicators between Dafne and PBI data.

    Args:
        dafne_data: Result of load_dafne_values().
        pbi_data:   Result of load_pbi_values().
        metrics:    Optional list of top-level metric keys to compare.
                    Defaults to ["output", "outcome", "impact"].

    Returns:
        {
            "output": {"dafne": ..., "pbi": ..., "delta": ..., "match": bool},
            "outcome": {...},
            "impact": {...},
            "discrepancies": [{"metric": str, "field": str, "dafne": ..., "pbi": ..., "delta": ...}],
            "summary": {"total_metrics": int, "matches": int, "mismatches": int}
        }

    # ES: Compara indicadores output/outcome/impact entre Dafne y PBI.
    # EN: Compares output/outcome/impact indicators between Dafne and PBI.
    """
    if metrics is None:
        metrics = ["output", "outcome", "impact"]

    discrepancies: list[dict] = []
    result: dict = {}
    total = 0
    matches = 0

    # ── Output ─────────────────────────────────────────────────────────────
    if "output" in metrics:
        d_out = dafne_data.get("output", {})
        p_out = pbi_data.get("output", {})
        output_result: dict = {}
        for field in ["area_ha", "beneficiarios", "organizaciones"]:
            d_val = d_out.get(field, 0) or 0
            p_val = p_out.get(field, 0) or 0
            delta = _delta(d_val, p_val)
            is_match = _values_match(d_val, p_val)
            output_result[field] = {
                "dafne": d_val,
                "pbi": p_val,
                "delta": delta,
                "match": is_match,
            }
            total += 1
            if is_match:
                matches += 1
            else:
                discrepancies.append({
                    "metric": "output",
                    "field": field,
                    "dafne": d_val,
                    "pbi": p_val,
                    "delta": delta,
                })
        result["output"] = output_result

    # ── Outcome ────────────────────────────────────────────────────────────
    if "outcome" in metrics:
        d_outcome = dafne_data.get("outcome", {})
        p_outcome = pbi_data.get("outcome", {})
        all_cdgs = set(d_outcome.keys()) | set(p_outcome.keys())
        outcome_result: dict = {}
        for cdg in sorted(all_cdgs):
            d_row = d_outcome.get(cdg, {})
            p_row = p_outcome.get(cdg, {})
            cdg_result: dict = {}
            for field in ["area_ha", "pct_logro"]:
                d_val = float(d_row.get(field, 0) or 0)
                p_val = float(p_row.get(field, 0) or 0)
                delta = _delta(d_val, p_val)
                is_match = _values_match(d_val, p_val)
                cdg_result[field] = {
                    "dafne": d_val,
                    "pbi": p_val,
                    "delta": delta,
                    "match": is_match,
                }
                total += 1
                if is_match:
                    matches += 1
                else:
                    discrepancies.append({
                        "metric": "outcome",
                        "field": f"{cdg}.{field}",
                        "cdg": cdg,
                        "dafne": d_val,
                        "pbi": p_val,
                        "delta": delta,
                    })
            outcome_result[cdg] = cdg_result
        result["outcome"] = outcome_result

    # ── Impact ─────────────────────────────────────────────────────────────
    if "impact" in metrics:
        d_imp = dafne_data.get("impact", {})
        p_imp = pbi_data.get("impact", {})
        d_val = float(d_imp.get("total_area_ha", 0) or 0)
        p_val = float(p_imp.get("total_area_ha", 0) or 0)
        delta = _delta(d_val, p_val)
        is_match = _values_match(d_val, p_val)
        result["impact"] = {
            "total_area_ha": {
                "dafne": d_val,
                "pbi": p_val,
                "delta": delta,
                "match": is_match,
            }
        }
        total += 1
        if is_match:
            matches += 1
        else:
            discrepancies.append({
                "metric": "impact",
                "field": "total_area_ha",
                "dafne": d_val,
                "pbi": p_val,
                "delta": delta,
            })

    result["discrepancies"] = discrepancies
    result["summary"] = {
        "total_metrics": total,
        "matches": matches,
        "mismatches": total - matches,
    }
    return result


def find_discrepancy_cause(
    dafne_row: dict,
    pbi_row: dict,
    cdg_actividad: str,
) -> dict:
    """
    Trace field-level discrepancy for a single CdgActvdd row.

    Compares every field present in either row and returns which fields differ.

    Returns:
        {
            "cdg":         str,
            "field_diffs": [{"field": str, "dafne": ..., "pbi": ...}],
            "match":       bool
        }

    # ES: Traza la causa de la discrepancia a nivel de campo para una actividad.
    # EN: Traces the field-level discrepancy cause for one activity row.
    """
    all_fields = set(dafne_row.keys()) | set(pbi_row.keys())
    diffs: list[dict] = []

    for field in sorted(all_fields):
        d_val = dafne_row.get(field)
        p_val = pbi_row.get(field)
        if not _values_match_raw(d_val, p_val):
            diffs.append({"field": field, "dafne": d_val, "pbi": p_val})

    return {
        "cdg": cdg_actividad,
        "field_diffs": diffs,
        "match": len(diffs) == 0,
    }


def decide_final_value(
    metric: str,
    dafne_val: float,
    pbi_val: float,
    decision: str,
    manual_val: Optional[float] = None,
    quarter: str = "",
) -> dict:
    """
    Record a final value decision for a metric.

    Args:
        metric:     Metric identifier (e.g. "output.area_ha").
        dafne_val:  Value from Dafne.
        pbi_val:    Value from Power BI.
        decision:   "dafne" | "pbi" | "manual".
        manual_val: Required when decision == "manual".
        quarter:    Quarter identifier for persisting to correct file.

    Returns:
        {"metric": str, "final_value": float, "decision": str, "recorded_at": str}

    # ES: Registra la decisión del valor final (Dafne, PBI o manual).
    # EN: Records the final value decision (Dafne, PBI, or manual).
    """
    if decision == "dafne":
        final = float(dafne_val)
    elif decision == "pbi":
        final = float(pbi_val)
    elif decision == "manual":
        if manual_val is None:
            raise ValueError("manual_val is required when decision == 'manual'")
        final = float(manual_val)
    else:
        raise ValueError(f"decision must be 'dafne', 'pbi', or 'manual', got: {decision}")

    entry = {
        "metric": metric,
        "dafne_val": float(dafne_val),
        "pbi_val": float(pbi_val),
        "decision": decision,
        "final_value": final,
        "manual_val": float(manual_val) if manual_val is not None else None,
        "quarter": quarter,
        "recorded_at": datetime.now().isoformat(timespec="seconds"),
    }

    _save_decision(quarter, metric, entry)
    return {
        "metric": metric,
        "final_value": final,
        "decision": decision,
        "recorded_at": entry["recorded_at"],
    }


def get_comparison_report(quarter: str) -> dict:
    """
    Load stored comparison results and decision history for a quarter.

    Returns:
        {
            "quarter":   str,
            "decisions": list[dict],
            "summary":   {"total_decisions": int, "dafne": int, "pbi": int, "manual": int}
        }

    # ES: Carga el reporte de comparación y decisiones del trimestre.
    # EN: Loads the comparison report and decisions for the quarter.
    """
    decisions = _load_decisions(quarter)
    summary = {
        "total_decisions": len(decisions),
        "dafne": sum(1 for d in decisions if d.get("decision") == "dafne"),
        "pbi": sum(1 for d in decisions if d.get("decision") == "pbi"),
        "manual": sum(1 for d in decisions if d.get("decision") == "manual"),
    }
    return {
        "quarter": quarter,
        "decisions": decisions,
        "summary": summary,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _col_idx(headers: list, candidates: list) -> Optional[int]:
    """Return first column index matching one of the candidate names."""
    for name in candidates:
        try:
            return headers.index(name)
        except ValueError:
            continue
    return None


def _delta(a: float, b: float) -> float:
    """Absolute difference rounded to 4 decimal places."""
    return round(abs(float(a) - float(b)), 4)


def _values_match(a, b) -> bool:
    """Numeric match within tolerance; non-numeric uses equality."""
    try:
        return abs(float(a) - float(b)) <= _TOLERANCE
    except (TypeError, ValueError):
        return a == b


def _values_match_raw(a, b) -> bool:
    """
    Match for raw cell values: numeric within tolerance, strings case-insensitive,
    None == None.
    """
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= _TOLERANCE
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def _decisions_file(quarter: str) -> pathlib.Path:
    os.makedirs(_COMPARE_DIR, exist_ok=True)
    return _COMPARE_DIR / f"{quarter}_decisions.json"


def _load_decisions(quarter: str) -> list[dict]:
    fp = _decisions_file(quarter)
    if not fp.exists():
        return []
    try:
        with open(fp, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def _save_decision(quarter: str, metric: str, entry: dict) -> None:
    """Upsert decision for metric in the quarter decisions file."""
    decisions = _load_decisions(quarter)
    # Replace existing entry for this metric if present
    decisions = [d for d in decisions if d.get("metric") != metric]
    decisions.append(entry)
    fp = _decisions_file(quarter)
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(decisions, f, ensure_ascii=False, indent=2)
