"""
AR Workflow Server – lightweight Flask app that gives non-technical
colleagues a browser UI for the full Altiplano Resiliente pipeline.

Start with:
    python app.py          (or: flask --app app run --port 5000)

Requires:
    pip install flask requests python-dotenv
"""
from __future__ import annotations

import csv
import io
import os
import pathlib
import re
import secrets
import subprocess
import textwrap
import time
import urllib.parse
from datetime import datetime

import requests
from dotenv import load_dotenv
from flask import Flask, jsonify, make_response, render_template, request, send_from_directory

from ar_utils import safe_resolve, smartsheet_request, stream_download_to_file
from paso3_criteria import (
    add_criterion, export_criteria_pdf, list_criteria_quarters,
    load_criteria, save_criteria,
)
from paso3_scripts import PASO3_SCRIPTS
from paso4_scripts import PASO4_SCRIPTS
from paso5_dafne import (
    get_integrado_status,
    get_reception_history,
    list_reception_quarters,
    place_dafne_file,
    receive_me_data,
    validate_integrado_xlsx,
)
from paso6_compare import (
    compare_me_values,
    decide_final_value,
    find_discrepancy_cause,
    get_comparison_report,
    load_dafne_values,
    load_pbi_values,
)
from paso6_powerbi import PASO6_SCRIPTS, check_publish_url, get_embed_url_via_api
from paso7_reports import generate_data_summary, generate_error_report, manage_backups
import paso1_agent
from orchestrator import PipelineOrchestrator
from pipeline_state import PipelineState

ROOT = pathlib.Path(__file__).parent

# Detect OneDrive roots BEFORE load_dotenv so that ${ONEDRIVE_*} placeholders
# inside .env are expanded by python-dotenv's built-in interpolation.
# / Detectar OneDrive ANTES de load_dotenv para que python-dotenv expanda
# los marcadores ${ONEDRIVE_*} del .env correctamente.
from env_paths import apply_onedrive_placeholders  # noqa: E402
apply_onedrive_placeholders()
load_dotenv(ROOT / ".env")

app = Flask(__name__, template_folder="templates", static_folder="static")

# Global singletons
_pipeline_state = PipelineState(ROOT / "downloads")
_orchestrator = PipelineOrchestrator(_pipeline_state)

# Local session token — protects sensitive endpoints from rogue localhost callers
_LOCAL_TOKEN = secrets.token_hex(16)

_PROTECTED_PREFIXES = (
    "/api/config/paths",
    "/api/pbi/launch",
    "/api/arcpy/open-pro",
    "/api/arcpy/close-pro",
    "/api/arcpy/run-add-to-map",
    "/api/arcpy/run-excel-to-point",
)


def _allowed_roots() -> list[str]:
    """Gather allowed filesystem roots from .env configuration."""
    roots = [str(ROOT / "downloads")]
    for key in ("WORKSPACE_PATH", "FOLDER_C1", "FOLDER_C2", "FOLDER_C3",
                "SMARTSHEET_ATTACH_DIR", "CSVPOINT_TO_GDB"):
        val = os.getenv(key, "")
        if val:
            roots.append(val)
    return roots


# ---------------------------------------------------------------------------
# Local auth — protect sensitive endpoints with a session token
# ---------------------------------------------------------------------------

@app.before_request
def _check_local_token():
    if not request.path.startswith(_PROTECTED_PREFIXES):
        return  # non-protected endpoint
    if request.method == "OPTIONS":
        return  # CORS preflight
    token = request.headers.get("X-Local-Token") or request.cookies.get("local_token")
    if token != _LOCAL_TOKEN:
        return jsonify(error="Unauthorized — invalid or missing local token"), 403


# ---------------------------------------------------------------------------
# Config helpers
# ---------------------------------------------------------------------------

# Variables that are local to each PC and editable from the web UI
_PATH_KEYS = [
    "FOLDER_C1", "FOLDER_C2", "FOLDER_C3",
    "SMARTSHEET_DATA_DIR", "SMARTSHEET_ATTACH_DIR", "EXCEL_DB_DIR",
    "WORKSPACE_PATH", "APRX", "AR_MAP_NAME",
    "CSVPOINT_TO_GDB",
    "ARCPY_PYTHON", "ARCGIS_PRO_VERSION",
]


def _cfg():
    return {
        "SMARTSHEET_TOKEN": os.getenv("SMARTSHEET_TOKEN") or os.getenv("TOKEN", ""),
        "SHEET_C1": os.getenv("SHEET_C1", ""),
        "SHEET_C2": os.getenv("SHEET_C2", ""),
        "SHEET_C3": os.getenv("SHEET_C3", ""),
        "FOLDER_C1": os.getenv("FOLDER_C1", ""),
        "FOLDER_C2": os.getenv("FOLDER_C2", ""),
        "FOLDER_C3": os.getenv("FOLDER_C3", ""),
        "DEFAULT_COMPONENT": os.getenv("DEFAULT_COMPONENT", "C1"),
        "ROW_START": os.getenv("ROW_START", "1"),
        "ROW_END": os.getenv("ROW_END", "9999"),
        "SMARTSHEET_DATA_DIR": os.getenv("SMARTSHEET_DATA_DIR") or os.getenv("SsheetDataDirectory", ""),
        "SMARTSHEET_ATTACH_DIR": os.getenv("SMARTSHEET_ATTACH_DIR") or os.getenv("ATTACH") or os.getenv("SsheetAttach", ""),
        "EXCEL_DB_DIR": os.getenv("EXCEL_DB_DIR") or os.getenv("EXCELDB", ""),
        "WORKSPACE_PATH": os.getenv("WORKSPACE_PATH") or os.getenv("workspace_path", ""),
        "APRX": os.getenv("APRX", ""),
        "AR_MAP_NAME": os.getenv("AR_MAP_NAME", "AR_EbA_Area"),
        "CSVPOINT_TO_GDB": os.getenv("CSVPOINT_TO_GDB") or os.getenv("GDB_PATH") or os.getenv("DEFAULT_GDB_PATH", ""),
        "ARCPY_PYTHON": os.getenv("ARCPY_PYTHON", ""),
        "ARCGIS_PRO_VERSION": os.getenv("ARCGIS_PRO_VERSION", ""),
        "AGOL_POLYGON_URL": os.getenv("AGOL_POLYGON_URL", ""),
        "AGOL_POINT_URL": os.getenv("AGOL_POINT_URL", ""),
    }


def _sheet_id(component: str) -> str:
    cfg = _cfg()
    return cfg.get(f"SHEET_{component}", "")


def _ss_headers():
    token = _cfg()["SMARTSHEET_TOKEN"]
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }


SS_BASE = "https://api.smartsheet.com/2.0"

# ---------------------------------------------------------------------------
# Smartsheet response cache — avoids redundant API calls within 5 minutes
# EN: Cache keyed by (component, sheet_id); entries expire after 300 seconds.
# ES: Caché por (componente, sheet_id); las entradas expiran en 300 segundos.
# ---------------------------------------------------------------------------
_SS_CACHE: dict = {}          # key -> {"ts": float, "data": dict}
_SS_CACHE_TTL = 300           # 5 minutes / 5 minutos

# ══ S3: Friendly error catalog ════════════════════════════════════════════
# NOTE: All user-facing text is in Spanish (end-user language).
# EN translations in comments for developer reference.

ERROR_MESSAGES = {
    # Smartsheet
    "ss_401": {
        # EN: "Invalid Smartsheet token"
        "title": "Token de Smartsheet inválido",
        # EN: "The Smartsheet access token is invalid or expired."
        "message": "El token de acceso a Smartsheet no es válido o ha expirado.",
        # EN: "Contact the system administrator to get a new token."
        "action": "Contacte al administrador del sistema para obtener un nuevo token.",
        "severity": "error",
    },
    "ss_403": {
        "title": "Sin permiso en Smartsheet",          # EN: No Smartsheet permission
        "message": "No tiene permisos para acceder a esta hoja.",
        "action": "Verifique que su cuenta tiene acceso a la hoja en Smartsheet.",
        "severity": "error",
    },
    "ss_404": {
        "title": "Hoja no encontrada",                 # EN: Sheet not found
        "message": "La hoja de Smartsheet no existe o fue eliminada.",
        "action": "Verifique el ID de la hoja en la configuración (.env).",
        "severity": "error",
    },
    "ss_429": {
        "title": "Demasiadas solicitudes",              # EN: Too many requests
        "message": "Se han enviado muchas solicitudes a Smartsheet en poco tiempo.",
        "action": "Espere un minuto y vuelva a intentar.",
        "severity": "warn",
    },
    "ss_network": {
        "title": "Sin conexión a Smartsheet",           # EN: No connection to Smartsheet
        "message": "No se pudo conectar con el servidor de Smartsheet.",
        "action": "Verifique su conexión a internet y vuelva a intentar.",
        "severity": "error",
    },
    # Configuration
    "no_token": {
        "title": "Token no configurado",                # EN: Token not configured
        "message": "No se ha configurado un token de acceso a Smartsheet.",
        "action": "Ejecute setup.bat o agregue SMARTSHEET_TOKEN al archivo .env.",
        "severity": "error",
    },
    "no_sheet_id": {
        "title": "ID de hoja no configurado",           # EN: Sheet ID not configured
        "message": "No se ha configurado el ID de la hoja para este componente.",
        "action": "Agregue SHEET_{{componente}} al archivo .env.",
        "severity": "error",
    },
    "no_folder": {
        "title": "Carpeta no configurada",              # EN: Folder not configured
        "message": "No se ha configurado la carpeta de destino.",
        "action": "Configure las rutas en Configuración → Rutas Locales.",
        "severity": "warn",
    },
    "folder_not_found": {
        "title": "Carpeta no encontrada",               # EN: Folder not found
        "message": "La carpeta '{path}' no existe en este equipo.",
        "action": "Verifique la ruta en Configuración → Rutas Locales, o cree la carpeta.",
        "severity": "error",
    },
    # Files
    "zip_invalid": {
        "title": "Archivo ZIP inválido",                # EN: Invalid ZIP file
        "message": "El archivo '{name}' no es un ZIP válido o está corrupto.",
        "action": "Pida al reportero que reenvíe el archivo.",
        "severity": "warn",
    },
    "shp_missing": {
        "title": "Shapefile incompleto",                # EN: Incomplete shapefile
        "message": "El archivo '{name}' no contiene todos los componentes necesarios (.shp, .shx, .dbf, .prj).",
        "action": "El shapefile debe incluir al menos 4 archivos. Pida al reportero que lo reenvíe completo.",
        "severity": "warn",
    },
    # GDB / ArcGIS
    "gdb_not_found": {
        "title": "Geodatabase no encontrada",           # EN: Geodatabase not found
        "message": "No se encontró la geodatabase en la ruta configurada.",
        "action": "Verifique CSVPOINT_TO_GDB en Configuración → Rutas Locales.",
        "severity": "error",
    },
    # ArcGIS Online
    "agol_auth": {
        "title": "Sin acceso a ArcGIS Online",          # EN: No AGOL access
        "message": "No se pudo autenticar con ArcGIS Online.",
        "action": (
            "Verifique las credenciales de ArcGIS en .env. "
            "Recuerde: la carga final requiere una cuenta interna de UICN (Guillermo/Marvin)."
        ),
        "severity": "error",
    },
    # Paso 0 — AGOL diagnostics
    "p0_no_agol_url": {
        # EN: No AGOL Feature Layer URL provided
        "title": "URL de Feature Layer no configurada",
        "message": "No se proporcionó la URL del Feature Layer de ArcGIS Online.",
        "action": "Configure AGOL_FEATURE_LAYER_URL en .env o ingrese la URL en el panel lateral.",
        "severity": "error",
    },
    "p0_agol_error": {
        # EN: AGOL query failed
        "title": "Error al consultar ArcGIS Online",
        "message": "No se pudo obtener datos del Feature Layer: {detail}",
        "action": "Verifique que la URL del Feature Layer es correcta y el servicio está accesible.",
        "severity": "error",
    },
    # Power BI
    "powerbi_no_license": {
        "title": "Licencia Power BI requerida",         # EN: Power BI license required
        "message": "Para actualizar el dashboard se necesita una licencia Power BI Pro.",
        "action": (
            "Solicite la licencia a través del portal de servicios de UICN, "
            "o utilice la licencia Pro existente de Dafne."
        ),
        "severity": "warn",
    },
    # ArcPy execution
    "arcpy_not_configured": {
        "title": "ArcPy no configurado",                # EN: ArcPy not configured
        "message": "No se encontró el ejecutable de Python de ArcGIS Pro.",
        "action": "Configure ARCPY_PYTHON en .env con la ruta al python.exe del entorno ArcGIS Pro.",
        "severity": "error",
    },
    "arcpy_timeout": {
        "title": "Script ArcPy agotó el tiempo",        # EN: ArcPy script timed out
        "message": "El script de ArcGIS Pro no terminó en el tiempo permitido (5 min).",
        "action": "Intente de nuevo. Si persiste, ejecute el script manualmente en ArcGIS Pro.",
        "severity": "error",
    },
    "arcpy_exec_error": {
        "title": "Error al ejecutar ArcPy",             # EN: ArcPy execution error
        "message": "No se pudo ejecutar el script: {detail}",
        "action": "Verifique la instalación de ArcGIS Pro y la ruta ARCPY_PYTHON en .env.",
        "severity": "error",
    },
    # Generic
    "network_error": {
        "title": "Error de conexión",                   # EN: Connection error
        "message": "No se pudo conectar con el servidor.",
        "action": "Verifique que el servidor está ejecutándose (ventana de run.bat abierta).",
        "severity": "error",
    },
    # Paso 1 Agent Loop
    "agent_no_dest": {
        "title": "Carpeta destino requerida",
        "message": "Debe especificar la carpeta destino (destFolder) para el agente.",
        "action": "Configure la carpeta destino en Paso 1b.",
        "severity": "error",
    },
    "agent_missing_params": {
        "title": "Parámetros insuficientes",
        "message": "Faltan parámetros requeridos para inicializar el agente.",
        "action": "Ejecute el diagnóstico primero y luego inicie el agente.",
        "severity": "error",
    },
    "agent_update_fail": {
        "title": "Error al actualizar agente",
        "message": "No se pudo actualizar el estado del agente: {detail}",
        "action": "Intente de nuevo. Si persiste, use el botón Reiniciar.",
        "severity": "error",
    },
    "unknown": {
        "title": "Error inesperado",                    # EN: Unexpected error
        "message": "Ocurrió un error inesperado.",
        "action": "Intente de nuevo. Si persiste, contacte al administrador.",
        "severity": "error",
    },
}


def friendly_error(error_code: str, **kwargs) -> dict:
    """Build a user-friendly error response (S3).

    Returns a dict with title/message/action in Spanish for UI display.
    EN: Build a structured error response using the catalog above.
    """
    template = ERROR_MESSAGES.get(error_code, ERROR_MESSAGES["unknown"])
    return {
        "error_code": error_code,
        "title": template["title"].format(**kwargs),
        "message": template["message"].format(**kwargs),
        "action": template["action"].format(**kwargs),
        "severity": template["severity"],
    }

# ══════════════════════════════════════════════════════════════════════════

# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    resp = make_response(render_template(
        "dashboard.html", cfg=_cfg(), local_token=_LOCAL_TOKEN,
    ))
    resp.set_cookie("local_token", _LOCAL_TOKEN, httponly=False, samesite="Strict")
    return resp


@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory("static", filename)


# ---------------------------------------------------------------------------
# API: Config
# ---------------------------------------------------------------------------

@app.route("/api/config")
def api_config():
    # Re-read .env so that manual edits outside the dashboard are picked up
    # without requiring a server restart.
    # / Releer .env para que cambios manuales se apliquen sin reiniciar.
    apply_onedrive_placeholders()
    load_dotenv(ROOT / ".env", override=True)
    cfg = _cfg()
    safe = {k: v for k, v in cfg.items() if k != "SMARTSHEET_TOKEN"}
    safe["has_token"] = bool(cfg["SMARTSHEET_TOKEN"])
    return jsonify(safe)


@app.route("/api/config/paths", methods=["GET"])
def api_config_paths_get():
    """Return current values of all PC-local path variables."""
    cfg = _cfg()
    return jsonify({k: cfg.get(k, "") for k in _PATH_KEYS})


@app.route("/api/config/paths", methods=["PUT"])
def api_config_paths_put():
    """Persist PC-local path variables to .env (never touches credentials)."""
    body = request.json or {}
    # Only allow known path keys — never credential or sheet-id fields
    unknown = set(body.keys()) - set(_PATH_KEYS)
    if unknown:
        return jsonify(error=f"Unknown or disallowed keys: {sorted(unknown)}"), 400

    # Reject path values containing traversal sequences after normalization
    for key, value in body.items():
        normalized = os.path.normpath(str(value))
        if ".." in normalized.split(os.sep):
            return jsonify(error=f"Path traversal not allowed in {key}"), 400

    try:
        from dotenv import set_key as _set_key
        env_path = str(ROOT / ".env")
        # Ensure .env exists
        if not (ROOT / ".env").exists():
            (ROOT / ".env").write_text("", encoding="utf-8")
        for key, value in body.items():
            _set_key(env_path, key, str(value), quote_mode="auto")
        # Reload env so the running process picks up new values immediately
        apply_onedrive_placeholders()
        load_dotenv(ROOT / ".env", override=True)
        return jsonify(saved=list(body.keys()))
    except Exception as exc:  # noqa: BLE001
        return jsonify(error=str(exc)), 500


# ---------------------------------------------------------------------------
# API: Smartsheet
# ---------------------------------------------------------------------------


def _apply_ss_filters(rows: list, body: dict) -> list:
    """Apply server-side row filters from request body.

    Supported filters (all optional):
    - rowStart / rowEnd: row-number range
    - filterTrimestres: list of quarter strings (TRIMESTRE QUE REPORTA)
    - filterTipos: list of grant-type strings; rows where NÚMERO DE CONTRATO
      contains the string (e.g. "PPD" matches "AR-PPD-001")
    - filterContratos: list of exact contract numbers (NÚMERO DE CONTRATO)
      (takes precedence over filterTipos when both are present)

    / Aplica filtros de filas del lado del servidor.
    """
    filtered = rows

    # Row-number range
    row_start = body.get("rowStart")
    row_end = body.get("rowEnd")
    if row_start is not None and row_end is not None:
        try:
            rs, re_ = int(row_start), int(row_end)
            filtered = [r for r in filtered if rs <= r["rowNumber"] <= re_]
        except (ValueError, TypeError):
            pass

    # Trimestre
    trimestres = body.get("filterTrimestres")
    if trimestres and isinstance(trimestres, list):
        t_set = set(trimestres)
        filtered = [
            r for r in filtered
            if (r["cells"].get("TRIMESTRE QUE REPORTA") or "") in t_set
        ]

    # Contrato takes precedence over Tipo
    contratos = body.get("filterContratos")
    tipos = body.get("filterTipos")
    if contratos and isinstance(contratos, list):
        c_set = set(contratos)
        filtered = [
            r for r in filtered
            if (r["cells"].get("NÚMERO DE CONTRATO") or "") in c_set
        ]
    elif tipos and isinstance(tipos, list):
        tipos_upper = [t.upper() for t in tipos]
        filtered = [
            r for r in filtered
            if any(
                t in (r["cells"].get("NÚMERO DE CONTRATO") or "").upper()
                for t in tipos_upper
            )
        ]

    return filtered

@app.route("/api/smartsheet/load", methods=["POST"])
def ss_load():
    body = request.json or {}
    comp = body.get("component", "C1")
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    # ── Cache check ──────────────────────────────────────────────────────────
    # EN: Return cached response if last fetch was within 5 minutes.
    # ES: Devuelve la respuesta en caché si la última consulta fue en los últimos 5 minutos.
    force_refresh = body.get("force_refresh", False)
    cache_key = f"{comp}:{sid}"
    cached = _SS_CACHE.get(cache_key)
    if not force_refresh and cached and (time.time() - cached["ts"]) < _SS_CACHE_TTL:
        cached_rows = cached["data"]["rows"]
        filtered_rows = _apply_ss_filters(cached_rows, body)
        payload = {
            "name": cached["data"]["name"],
            "totalRows": cached["data"]["totalRows"],
            "filteredRows": len(filtered_rows),
            "columns": cached["data"]["columns"],
            "rows": filtered_rows,
            "from_cache": True,
            "cache_age_seconds": int(time.time() - cached["ts"]),
        }
        return jsonify(**payload)
    # ─────────────────────────────────────────────────────────────────────────

    try:
        r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(),
            params={"include": "attachments,objectValue"},
            timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if r.status_code == 403:
        return jsonify(friendly_error("ss_403")), 403
    if r.status_code == 404:
        return jsonify(friendly_error("ss_404")), 404
    if r.status_code == 429:
        return jsonify(friendly_error("ss_429")), 429
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    data = r.json()

    # Build column map
    col_map = {c["title"]: c["id"] for c in data.get("columns", [])}
    # Summarise rows (comment: None by default — use /api/smartsheet/comments to load)
    rows_out = []
    for row in data.get("rows", []):
        cells = {}
        for cell in row.get("cells", []):
            for title, cid in col_map.items():
                if cell.get("columnId") == cid:
                    # Prefer numeric value; fall back to objectValue then displayValue
                    # (formula cells and some column types omit 'value')
                    v = cell.get("value")
                    if v is None:
                        obj = cell.get("objectValue")
                        if isinstance(obj, dict):
                            v = obj.get("value") or obj.get("number")
                        elif obj is not None:
                            v = obj
                    if v is None:
                        v = cell.get("displayValue")
                    cells[title] = v
        rows_out.append({
            "rowNumber": row.get("rowNumber"),
            "id": row.get("id"),
            "cells": cells,
            "attachments": [
                {"id": a["id"], "name": a.get("name", "")}
                for a in row.get("attachments", [])
            ],
            "comment": None,
        })

    total_unfiltered = len(rows_out)
    result = {
        "name": data.get("name"),
        "permalink": data.get("permalink", ""),
        "totalRows": total_unfiltered,
        "columns": list(col_map.keys()),
        "rows": rows_out,
    }
    # Store full sheet in cache; row slicing happens after cache lookup
    _SS_CACHE[cache_key] = {"ts": time.time(), "data": result}

    # ── Server-side filtering (applied after cache store) ───────────────────
    filtered_rows = _apply_ss_filters(rows_out, body)

    response = {
        "name": result["name"],
        "totalRows": total_unfiltered,
        "filteredRows": len(filtered_rows),
        "columns": result["columns"],
        "rows": filtered_rows,
    }
    return jsonify(**response, from_cache=False)


@app.route("/api/smartsheet/comments", methods=["POST"])
def ss_comments():
    """Lazy-load row comments for a sheet component.
    / Carga diferida de comentarios por fila de un componente.
    """
    body = request.json or {}
    comp = body.get("component", "C1")
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    row_comments = {}
    try:
        dr = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}/discussions",
            headers=_ss_headers(),
            params={"include": "comments"},
            timeout=30,
        )
        if dr.ok:
            for disc in dr.json().get("data", []):
                if disc.get("parentType") != "ROW":
                    continue
                parent_id = disc.get("parentId")
                if parent_id in row_comments:
                    continue
                for cmt in disc.get("comments", []):
                    row_comments[parent_id] = {
                        "text": cmt.get("text", ""),
                        "author": (cmt.get("createdBy") or {}).get("name", ""),
                        "date": (cmt.get("createdAt") or "")[:10],
                    }
                    break
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503

    return jsonify(comments=row_comments)


@app.route("/api/smartsheet/diagnose", methods=["POST"])
def ss_diagnose():
    """Return raw cell data for the ha column of the first 10 data rows (rows 6-15).
    Used to diagnose why 'TOTAL DE HECTÁREAS' shows 0 in the summary.
    """
    body = request.json or {}
    comp = body.get("component", "C1")
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(error="no sheet id"), 400
    try:
        r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(),
            params={"include": "attachments,objectValue"},
            timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(error="network"), 503
    if not r.ok:
        return jsonify(error=r.status_code, text=r.text[:500]), r.status_code
    data = r.json()

    # Find ha column id
    ha_col_id = None
    ha_col_name = None
    cols_info = []
    for c in data.get("columns", []):
        cols_info.append({"title": c["title"], "id": c["id"], "type": c.get("type")})
        if "HECT" in c["title"].upper() or "HA" in c["title"].upper():
            ha_col_id = c["id"]
            ha_col_name = c["title"]

    # Sample first 15 rows' raw cells for ha column
    sample = []
    for row in data.get("rows", [])[:15]:
        for cell in row.get("cells", []):
            if cell.get("columnId") == ha_col_id:
                sample.append({
                    "row": row.get("rowNumber"),
                    "value": cell.get("value"),
                    "displayValue": cell.get("displayValue"),
                    "objectValue": cell.get("objectValue"),
                    "formula": cell.get("formula"),
                })
    return jsonify(
        ha_col_name=ha_col_name,
        ha_col_id=ha_col_id,
        columns=cols_info,
        ha_cells_sample=sample,
    )


@app.route("/api/smartsheet/cache/clear", methods=["POST"])
def ss_cache_clear():
    """Clear the in-memory Smartsheet cache (forces fresh API fetch on next load)."""
    body = request.json or {}
    comp = body.get("component")  # optional: clear only one component
    cleared = []
    keys = list(_SS_CACHE.keys())
    for k in keys:
        if comp is None or k.startswith(comp + ":"):
            del _SS_CACHE[k]
            cleared.append(k)
    return jsonify(cleared=cleared)


@app.route("/api/smartsheet/generate-codes", methods=["POST"])
def ss_generate_codes():
    """Generate missing CÓDIGO DE LA ACTIVIDAD values."""
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    row_ids = body.get("rowIds")          # optional list of specific Smartsheet row IDs
    if row_ids is not None and len(row_ids) == 0:
        return jsonify(error="rowIds was provided but empty — specify target rows or omit the field"), 400
    row_ids_set = set(row_ids) if row_ids else None
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    fecha_id = col_map.get("FECHA DE LA ACTIVIDAD")
    code_id = col_map.get("CÓDIGO DE LA ACTIVIDAD")
    who_id = col_map.get("NOMBRE DE QUIEN REPORTA") if comp != "C2" else col_map.get("ORGANIZACIÓN")

    if not all([fecha_id, code_id, who_id]):
        return jsonify(friendly_error("unknown")), 400

    def cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    existing = set()
    for row in sheet["rows"]:
        v = cell_val(row, code_id)
        if v:
            existing.add(v)

    import itertools

    # Pre-filter rows by range (and C2 filters when no explicit row IDs)
    candidate_rows = sheet["rows"]
    if row_ids_set is not None:
        candidate_rows = [row for row in candidate_rows if row.get("id") in row_ids_set]
    else:
        candidate_rows = [
            row for row in candidate_rows
            if row_start <= row.get("rowNumber", 0) <= row_end
        ]
        if comp == "C2":
            candidate_rows = _c2_filter_rows(candidate_rows, body, col_map, cell_val)

    updates = []
    skipped = {"already_has_code": 0, "missing_date": 0, "missing_who": 0}
    who_col_name = "NOMBRE DE QUIEN REPORTA" if comp != "C2" else "ORGANIZACIÓN"
    for row in candidate_rows:
        if cell_val(row, code_id):
            skipped["already_has_code"] += 1
            continue
        fecha = cell_val(row, fecha_id)
        who_raw = cell_val(row, who_id) or ""
        who = who_raw[:3] if comp != "C2" else who_raw
        if not fecha:
            skipped["missing_date"] += 1
            continue
        if not who:
            skipped["missing_who"] += 1
            continue
        base_date = str(fecha).replace("-", "")[2:]
        base = f"{base_date}_{comp}_{who}_"
        code = None
        for length in range(1, 5):
            for combo in itertools.product("abcdefghijklmnopqrstuvwxyz", repeat=length):
                candidate = base + "".join(combo)
                if candidate not in existing:
                    existing.add(candidate)
                    code = candidate
                    break
            if code:
                break
        if code:
            updates.append({"id": row["id"], "cells": [{"columnId": code_id, "value": code}]})

    # Build row_id -> code mapping for client-side cache patching
    code_col_name = "CÓDIGO DE LA ACTIVIDAD"
    row_patches = {u["id"]: {code_col_name: u["cells"][0]["value"]} for u in updates}

    updated = 0
    batch_results = []
    sent_ids = set()
    batch_size = 50
    for i in range(0, len(updates), batch_size):
        batch = updates[i:i + batch_size]
        batch_num = i // batch_size + 1
        r = smartsheet_request(
            "PUT",
            f"{SS_BASE}/sheets/{sid}/rows",
            headers=_ss_headers(),
            json=batch,
            timeout=30,
        )
        if r.ok:
            updated += len(batch)
            sent_ids.update(u["id"] for u in batch)
            batch_results.append({"batch": batch_num, "status": "ok"})
        else:
            batch_results.append({"batch": batch_num, "status": "error", "code": r.status_code, "message": r.text[:200]})

    # Only return patches for rows that were actually sent successfully
    patches = [{"rowId": rid, "cells": cells}
               for rid, cells in row_patches.items() if rid in sent_ids]

    # Invalidate server cache so next full load gets fresh data
    cache_key = f"{comp}:{sid}"
    _SS_CACHE.pop(cache_key, None)

    return jsonify(
        generated=updated,
        total_rows_checked=len(candidate_rows),
        batch_results=batch_results,
        skipped=skipped,
        who_column=who_col_name,
        patches=patches,
    )


@app.route("/api/smartsheet/fix-dates", methods=["POST"])
def ss_fix_dates():
    """Fix text-format dates in FECHA DE LA ACTIVIDAD column.

    Receives a list of {rowId, date} pairs where date is already in ISO format
    (YYYY-MM-DD) and updates the Smartsheet cells.
    """
    import re as _re

    body = request.json or {}
    comp = body.get("component", "C1")
    fixes = body.get("fixes", [])
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400
    if not fixes:
        return jsonify(updated=0, patches=[])

    # Validate ISO date format
    iso_re = _re.compile(r"^\d{4}-\d{2}-\d{2}$")
    valid_fixes = [f for f in fixes if isinstance(f.get("rowId"), (int, float)) and iso_re.match(str(f.get("date", "")))]
    if not valid_fixes:
        return jsonify(friendly_error("unknown")), 400

    # Get column ID for FECHA DE LA ACTIVIDAD
    try:
        r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(),
            params={"include": "objectValue"},
            timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    data = r.json()
    col_map = {c["title"]: c["id"] for c in data.get("columns", [])}
    fecha_id = col_map.get("FECHA DE LA ACTIVIDAD")
    if not fecha_id:
        return jsonify(friendly_error("unknown")), 400

    updates = [
        {"id": int(f["rowId"]), "cells": [{"columnId": fecha_id, "value": f["date"]}]}
        for f in valid_fixes
    ]

    updated = 0
    batch_results = []
    patches = []
    for i in range(0, len(updates), 100):
        batch = updates[i : i + 100]
        resp = smartsheet_request(
            "PUT",
            f"{SS_BASE}/sheets/{sid}/rows",
            headers=_ss_headers(),
            json=batch,
            timeout=30,
        )
        batch_results.append({"status": resp.status_code, "count": len(batch)})
        if resp.ok:
            updated += len(batch)
            for u in batch:
                patches.append({"rowId": u["id"], "cells": {"FECHA DE LA ACTIVIDAD": u["cells"][0]["value"]}})

    # Invalidate cache
    cache_key = f"{comp}:{sid}"
    _SS_CACHE.pop(cache_key, None)

    return jsonify(updated=updated, total=len(updates), batch_results=batch_results, patches=patches)


@app.route("/api/smartsheet/update-review", methods=["POST"])
def ss_update_review():
    """Update Calidad SIG: Espera if row has shapefile ZIP, Sí otherwise."""
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400
    if comp == "C3":
        return jsonify(friendly_error("unknown")), 400  # EN: C3 has no Calidad SIG column

    # Include attachments so we can detect shapefile ZIPs
    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(),
                               params={"include": "attachments"}, timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    code_id = col_map.get("CÓDIGO DE LA ACTIVIDAD")
    sig_id = col_map.get("Calidad SIG")

    if not sig_id:
        return jsonify(friendly_error("unknown")), 400  # EN: Calidad SIG column not found

    # Manual quality ratings that must not be overwritten automatically
    MANUAL_RATINGS = {"Rojo", "Amarillo", "Verde"}
    SHAPE_KEYWORDS = ("shp", "shape", "area", "área")

    def cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    def has_shapefile_zip(row):
        for a in row.get("attachments", []):
            n = a.get("name", "").lower()
            if n.endswith(".zip") and any(kw in n for kw in SHAPE_KEYWORDS):
                return True
        return False

    # Filter rows by range and C2 filters
    target_rows = [
        row for row in sheet["rows"]
        if row_start <= row.get("rowNumber", 0) <= row_end
    ]
    # EN: If frontend sent filtered rowNumbers, restrict to those rows only
    # ES: Si el frontend envió rowNumbers filtrados, restringir solo a esas filas
    row_numbers = body.get("rowNumbers")
    if row_numbers:
        rn_set = set(row_numbers)
        target_rows = [r for r in target_rows if r.get("rowNumber") in rn_set]
    if comp == "C2":
        target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val)

    updates = []
    # Diagnostics: track why rows are skipped / Diagnóstico: rastrear por qué se omiten filas
    diag_no_code = 0        # rows without CÓDIGO DE LA ACTIVIDAD
    diag_no_code_cleared = 0  # rows without code whose Calidad SIG was cleared to null
    diag_manual = 0         # rows with manual rating (Rojo/Amarillo/Verde)
    diag_already_ok = 0     # rows already set to correct value
    diag_already_detail = {"Espera": 0, "Sí": 0}
    diag_with_shape = 0     # rows that have shapefile ZIP
    diag_total = len(target_rows)
    for row in target_rows:
        has_zip = has_shapefile_zip(row)
        if has_zip:
            diag_with_shape += 1
        if not cell_val(row, code_id):
            diag_no_code += 1
            # EN: Rows without CÓDIGO should have Calidad SIG = null; clear if set
            # ES: Filas sin CÓDIGO deben tener Calidad SIG = null; limpiar si tiene valor
            current = cell_val(row, sig_id)
            if current is not None and str(current).strip() != "":
                updates.append({"id": row["id"], "cells": [{"columnId": sig_id, "value": None}]})
                diag_no_code_cleared += 1
            continue
        current = cell_val(row, sig_id)
        if current in MANUAL_RATINGS:
            diag_manual += 1
            continue  # preserve manual quality ratings
        new_val = "Espera" if has_zip else "Sí"
        if current == new_val:
            diag_already_ok += 1
            diag_already_detail[current] = diag_already_detail.get(current, 0) + 1
            continue  # already correct
        updates.append({"id": row["id"], "cells": [{"columnId": sig_id, "value": new_val}]})

    espera_count = sum(1 for u in updates if u["cells"][0]["value"] == "Espera")
    si_count = sum(1 for u in updates if u["cells"][0]["value"] == "Sí")

    diagnostics = {
        "totalRows": diag_total,
        "noCode": diag_no_code,
        "noCodeCleared": diag_no_code_cleared,
        "manualRating": diag_manual,
        "alreadyCorrect": diag_already_ok,
        "alreadyDetail": diag_already_detail,
        "withShapefile": diag_with_shape,
        "toUpdate": len(updates),
    }

    # Build row_id -> new Calidad SIG value for patches
    sig_col_name = "Calidad SIG"
    row_patches = {u["id"]: {sig_col_name: u["cells"][0]["value"]} for u in updates}

    updated = 0
    batch_results = []
    sent_ids = set()
    for i in range(0, len(updates), 50):
        batch = updates[i:i + 50]
        batch_num = i // 50 + 1
        r = smartsheet_request("PUT", f"{SS_BASE}/sheets/{sid}/rows", headers=_ss_headers(), json=batch, timeout=30)
        if r.ok:
            updated += len(batch)
            sent_ids.update(u["id"] for u in batch)
            batch_results.append({"batch": batch_num, "status": "ok"})
        else:
            batch_results.append({"batch": batch_num, "status": "error", "code": r.status_code, "message": r.text[:200]})

    patches = [{"rowId": rid, "cells": cells}
               for rid, cells in row_patches.items() if rid in sent_ids]

    # Invalidate server cache
    cache_key = f"{comp}:{sid}"
    _SS_CACHE.pop(cache_key, None)

    return jsonify(updated=updated, espera=espera_count, si=si_count,
                   batch_results=batch_results, patches=patches,
                   diagnostics=diagnostics)


@app.route("/api/smartsheet/set-calidad", methods=["POST"])
def ss_set_calidad():
    """Bulk-set Calidad SIG to a given value for rows matching criteria."""
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    new_val = body.get("value", "")  # "Espera", "Sí", or "" (clear)
    # Which rows to target: "shape" = rows with shapefile, "all" = all rows with code
    target = body.get("target", "shape")
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400
    if comp == "C3":
        return jsonify(friendly_error("unknown")), 400

    # When specific row numbers are provided, pass them to the Smartsheet API
    # to avoid fetching all rows (which can lose attachment data on large sheets).
    allowed_rows = body.get("rowNumbers")
    api_params = {"include": "attachments"}
    if allowed_rows:
        api_params["rowNumbers"] = ",".join(str(rn) for rn in allowed_rows)

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(),
                               params=api_params, timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    code_id = col_map.get("CÓDIGO DE LA ACTIVIDAD")
    sig_id = col_map.get("Calidad SIG")
    if not sig_id:
        return jsonify(friendly_error("unknown")), 400

    SHAPE_KEYWORDS = ("shp", "shape", "area", "área")

    def cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    def has_shapefile_zip(row):
        for a in row.get("attachments", []):
            n = a.get("name", "").lower()
            if n.endswith(".zip") and any(kw in n for kw in SHAPE_KEYWORDS):
                return True
        return False

    if allowed_rows:
        # API already filtered by rowNumbers; use all returned rows
        target_rows = sheet.get("rows", [])
    else:
        target_rows = [
            row for row in sheet["rows"]
            if row_start <= row.get("rowNumber", 0) <= row_end
        ]
        if comp == "C2":
            target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val)

    # Diagnostics: track why rows are skipped / Diagnóstico: rastrear por qué se omiten filas
    diag_total = len(target_rows)
    diag_no_code = 0
    diag_no_shape = 0
    diag_already = 0

    updates = []
    sig_col_name = "Calidad SIG"
    for row in target_rows:
        if not cell_val(row, code_id):
            diag_no_code += 1
            continue
        if target == "shape" and not has_shapefile_zip(row):
            diag_no_shape += 1
            continue
        current = cell_val(row, sig_id)
        if current == new_val:
            diag_already += 1
            continue
        updates.append({"id": row["id"], "cells": [{"columnId": sig_id, "value": new_val}]})

    row_patches = {u["id"]: {sig_col_name: u["cells"][0]["value"]} for u in updates}
    updated = 0
    sent_ids = set()
    for i in range(0, len(updates), 50):
        batch = updates[i:i + 50]
        r = smartsheet_request("PUT", f"{SS_BASE}/sheets/{sid}/rows", headers=_ss_headers(), json=batch, timeout=30)
        if r.ok:
            updated += len(batch)
            sent_ids.update(u["id"] for u in batch)

    patches = [{"rowId": rid, "cells": cells}
               for rid, cells in row_patches.items() if rid in sent_ids]

    cache_key = f"{comp}:{sid}"
    _SS_CACHE.pop(cache_key, None)

    diagnostics = {
        "totalRows": diag_total,
        "noCode": diag_no_code,
        "noShapefile": diag_no_shape,
        "alreadyCorrect": diag_already,
    }

    return jsonify(updated=updated, value=new_val, patches=patches, diagnostics=diagnostics)


def _c2_filter_rows(rows, body, col_map, cell_val_fn):
    """Apply C2 grant-type and/or contrato filters to a list of Smartsheet rows.

    NÚMERO DE CONTRATO is set only on parent/group-header rows.
    Child activity rows inherit the contrato from the nearest preceding parent.
    """
    grant_type = body.get("c2GrantType", "")
    contrato_filter = body.get("c2Contrato", "")
    contrato_id = col_map.get("NÚMERO DE CONTRATO")
    if not grant_type and not contrato_filter:
        return rows
    filtered = []
    current_contrato = ""
    for row in rows:
        if not contrato_id:
            filtered.append(row)
            continue
        contrato = cell_val_fn(row, contrato_id) or ""
        if contrato:
            current_contrato = contrato
        if contrato_filter:
            if current_contrato != contrato_filter:
                continue
        elif grant_type:
            if grant_type not in str(current_contrato):
                continue
        filtered.append(row)
    return filtered


# ---------------------------------------------------------------------------
# Canonical ABE values — must match _ABE_ABBR_MAP in ar_utils.py
# Valores AbE canónicos — deben coincidir con _ABE_ABBR_MAP en ar_utils.py
# ---------------------------------------------------------------------------
_ABE_CANONICAL = {
    "Sistema Agroforestal | Cultivos anuales",
    "Sistema Agroforestal | Cultivos perennes",
    "Sistema Silvopastoril",
    "Plantaciones Forestales",
    "Bosque Natural con Fines de Producción",
    "Bosque Natural con Fines de Protección",
    "Reforestación con Fines de Restauración",
    "Sistema Agroforestal | Conservación de suelo y agua",
}


def _fix_abe_value(raw: str) -> str | None:
    """Return the canonical ABE string for a known fixable variant, or None if already canonical / unknown.

    EN: Only handles variants that can be auto-corrected without human review:
        - wrong capitalisation/accents on a known value
        - reversed order (Conservación de suelo y agua | Sistema Agroforestal)
        - truncated strings where the canonical value can be inferred
    ES: Solo corrige variantes que pueden ser auto-corregidas sin revisión humana.
    """
    if not raw or not isinstance(raw, str):
        return None
    if raw in _ABE_CANONICAL:
        return None  # already correct / ya es correcto
    import unicodedata as _ud

    def _norm(s: str) -> str:
        return _ud.normalize("NFD", s).encode("ASCII", "ignore").decode().lower().strip()

    raw_norm = _norm(raw)

    # Map of normalised canonical → canonical display
    _norm_to_canon = {_norm(c): c for c in _ABE_CANONICAL}

    # 1. Exact normalised match
    if raw_norm in _norm_to_canon:
        return _norm_to_canon[raw_norm]

    # 2. Partial/prefix match for truncated variants (min 15 chars to avoid false positives)
    if len(raw_norm) >= 15:
        for k, canon in _norm_to_canon.items():
            if raw_norm in k or k.startswith(raw_norm[:20]):
                return canon

    # 3. "suelos" misspelling → suelo y agua canonical
    if "suelos" in raw_norm and "agroforestal" in raw_norm:
        return "Sistema Agroforestal | Conservación de suelo y agua"

    return None


@app.route("/api/smartsheet/fix-abe", methods=["POST"])
def ss_fix_abe():
    """Correct fixable ABE value variants directly in Smartsheet.

    Accepts two modes:
    1. Targeted (preferred): { component, old_value, new_value }
       EN: Replaces ALL occurrences of old_value with new_value (user-confirmed via dialog).
       ES: Reemplaza TODAS las ocurrencias de old_value con new_value (confirmado por el usuario).
    2. Auto-scan (legacy): { component, rowStart?, rowEnd? }
       EN: Scans and auto-corrects known mis-accented/truncated variants.
       ES: Escanea y corrige automáticamente variantes con acentos incorrectos/truncadas.
    """
    body = request.json or {}
    comp = body.get("component", "C1")
    # EN: Targeted mode — user confirmed old→new via dialog
    old_value = body.get("old_value")
    new_value = body.get("new_value")
    row_start = int(body.get("rowStart", 1))
    row_end = int(body.get("rowEnd", 999999))
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    # EN: Validate new_value must be a recognised canonical ABE value
    # ES: Validar que new_value sea un valor AbE canónico reconocido
    if old_value is not None and new_value is not None:
        if new_value not in _ABE_CANONICAL:
            return jsonify(friendly_error("unknown")), 400

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code

    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    abe_col_id = col_map.get("ACCIONES DE RESTAURACIÓN AbE")
    code_col_id = col_map.get("CÓDIGO DE LA ACTIVIDAD")
    if not abe_col_id:
        return jsonify(friendly_error("unknown")), 400

    def _cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    updates = []
    skipped_unknown = []
    already_ok = 0

    for row in sheet.get("rows", []):
        rn = row.get("rowNumber", 0)
        if rn < row_start or rn > row_end:
            continue
        raw_abe = _cell_val(row, abe_col_id)
        if not raw_abe:
            continue

        if old_value is not None:
            # EN: Targeted mode: only fix rows whose ABE exactly matches old_value
            # ES: Modo dirigido: solo corregir filas cuyo AbE coincide exactamente con old_value
            if raw_abe == old_value:
                updates.append({
                    "id": row["id"], "rowNumber": rn,
                    "oldValue": raw_abe, "newValue": new_value,
                    "cells": [{"columnId": abe_col_id, "value": new_value}],
                })
        else:
            # EN: Auto-scan mode: infer the canonical value
            # ES: Modo auto-escaneo: inferir el valor canónico
            fixed = _fix_abe_value(raw_abe)
            if fixed is None:
                if raw_abe in _ABE_CANONICAL:
                    already_ok += 1
                else:
                    code = _cell_val(row, code_col_id) or f"fila {rn}"
                    skipped_unknown.append({"rowNumber": rn, "codigo": code, "value": raw_abe})
                continue
            updates.append({
                "id": row["id"], "rowNumber": rn,
                "oldValue": raw_abe, "newValue": fixed,
                "cells": [{"columnId": abe_col_id, "value": fixed}],
            })

    if not updates:
        return jsonify(
            updated=0, already_ok=already_ok,
            skipped_unknown=skipped_unknown,
            message="No se encontraron valores AbE corregibles. / No fixable AbE values found."
        )

    # Send in batches of 50
    payload = [{"id": u["id"], "cells": u["cells"]} for u in updates]
    total_updated = 0
    batch_results = []
    for i in range(0, len(payload), 50):
        batch = payload[i:i + 50]
        br = smartsheet_request("PUT", f"{SS_BASE}/sheets/{sid}/rows",
                                headers=_ss_headers(), json=batch, timeout=30)
        bn = i // 50 + 1
        if br.ok:
            total_updated += len(batch)
            batch_results.append({"batch": bn, "status": "ok", "count": len(batch)})
        else:
            batch_results.append({"batch": bn, "status": "error",
                                   "code": br.status_code, "message": br.text[:300]})

    # Invalidate server cache
    _SS_CACHE.pop(f"{comp}:{sid}", None)

    fixed_rows = [{"rowNumber": u["rowNumber"], "codigo": None, "oldValue": u["oldValue"], "newValue": u["newValue"]}
                  for u in updates]

    return jsonify(
        updated=total_updated,
        already_ok=already_ok,
        skipped_unknown=skipped_unknown,
        fixed_rows=fixed_rows,
        batch_results=batch_results,
    )


@app.route("/api/smartsheet/fill-down", methods=["POST"])
def ss_fill_down():
    """Fill empty cells downward for specified columns within a row range."""
    body = request.json or {}
    comp = body.get("component", "C2")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    columns_to_fill = body.get("columns", [])
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    def cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    # Get rows in range
    target_rows = [
        row for row in sheet.get("rows", [])
        if row_start <= row.get("rowNumber", 0) <= row_end
    ]

    # Apply C2 grant-type filter if provided
    target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val)

    # Collect updates: fill down within same NÚMERO DE CONTRATO group,
    # only for rows that have FECHA DE LA ACTIVIDAD
    contrato_id = col_map.get("NÚMERO DE CONTRATO")
    fecha_id = col_map.get("FECHA DE LA ACTIVIDAD")
    row_updates = {}  # row_id -> {col_id: value, ...}
    for col_name in columns_to_fill:
        cid = col_map.get(col_name)
        if not cid:
            continue
        last_contrato = None
        last_value = None
        for row in target_rows:
            # Reset when contract changes
            cur_contrato = cell_val(row, contrato_id) if contrato_id else None
            if cur_contrato and cur_contrato != last_contrato:
                last_contrato = cur_contrato
                last_value = None
            current = cell_val(row, cid)
            if current:
                last_value = current
            elif last_value and (cell_val(row, fecha_id) if fecha_id else True):
                rid = row["id"]
                if rid not in row_updates:
                    row_updates[rid] = {}
                row_updates[rid][cid] = last_value

    updates = [
        {"id": rid, "cells": [{"columnId": cid, "value": val} for cid, val in cols.items()]}
        for rid, cols in row_updates.items()
    ]

    # Build reverse map: column_id -> column_name for patches
    id_to_name = {v: k for k, v in col_map.items()}

    updated = 0
    batch_results = []
    sent_ids = set()
    for i in range(0, len(updates), 50):
        batch = updates[i:i + 50]
        batch_num = i // 50 + 1
        r = smartsheet_request(
            "PUT",
            f"{SS_BASE}/sheets/{sid}/rows",
            headers=_ss_headers(),
            json=batch,
            timeout=30,
        )
        if r.ok:
            updated += len(batch)
            sent_ids.update(u["id"] for u in batch)
            batch_results.append({"batch": batch_num, "status": "ok"})
        else:
            batch_results.append({"batch": batch_num, "status": "error", "code": r.status_code, "message": r.text[:200]})

    # Return patches keyed by column name for client-side cache patching
    patches = [
        {"rowId": rid, "cells": {id_to_name.get(cid, str(cid)): val for cid, val in cols.items()}}
        for rid, cols in row_updates.items() if rid in sent_ids
    ]

    # Invalidate server cache
    cache_key = f"{comp}:{sid}"
    _SS_CACHE.pop(cache_key, None)

    return jsonify(updated=updated, total_checked=len(target_rows), batch_results=batch_results, patches=patches)


@app.route("/api/smartsheet/fill-down-preview", methods=["POST"])
def ss_fill_down_preview():
    """Preview which empty cells can be filled down (dry-run, no writes).

    Returns per-column counts and per-row details of what would be filled,
    so the UI can show a confirmation dialog before actually writing.
    """
    body = request.json or {}
    comp = body.get("component", "C2")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    columns_to_check = body.get("columns", [])
    row_ids = body.get("rowIds")
    row_ids_set = set(row_ids) if row_ids else None
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    def cell_val(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    # Select target rows (same logic as fill-down)
    target_rows = sheet.get("rows", [])
    if row_ids_set is not None:
        target_rows = [row for row in target_rows if row.get("id") in row_ids_set]
    else:
        target_rows = [
            row for row in target_rows
            if row_start <= row.get("rowNumber", 0) <= row_end
        ]
        if comp == "C2":
            target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val)

    contrato_id = col_map.get("NÚMERO DE CONTRATO")
    fecha_id = col_map.get("FECHA DE LA ACTIVIDAD")

    # Build preview: which cells would be filled and with what value
    fillable = []  # [{rowNumber, rowId, column, value}, ...]
    col_counts = {}  # column_name -> count of fillable cells
    for col_name in columns_to_check:
        cid = col_map.get(col_name)
        if not cid:
            continue
        last_contrato = None
        last_value = None
        for row in target_rows:
            cur_contrato = cell_val(row, contrato_id) if contrato_id else None
            if cur_contrato and cur_contrato != last_contrato:
                last_contrato = cur_contrato
                last_value = None
            current = cell_val(row, cid)
            if current:
                last_value = current
            elif last_value and (cell_val(row, fecha_id) if fecha_id else True):
                fillable.append({
                    "rowNumber": row.get("rowNumber"),
                    "rowId": row["id"],
                    "column": col_name,
                    "value": last_value,
                })
                col_counts[col_name] = col_counts.get(col_name, 0) + 1

    # Also report cells that are empty but NOT fillable (no upper value in group)
    unfillable = 0
    for col_name in columns_to_check:
        cid = col_map.get(col_name)
        if not cid:
            continue
        last_contrato = None
        last_value = None
        for row in target_rows:
            cur_contrato = cell_val(row, contrato_id) if contrato_id else None
            if cur_contrato and cur_contrato != last_contrato:
                last_contrato = cur_contrato
                last_value = None
            current = cell_val(row, cid)
            if current:
                last_value = current
            elif not last_value and (cell_val(row, fecha_id) if fecha_id else True):
                unfillable += 1

    return jsonify(
        fillable=fillable,
        col_counts=col_counts,
        total_fillable=len(fillable),
        unfillable=unfillable,
    )


@app.route("/api/smartsheet/attachments", methods=["POST"])
def ss_attachments():
    """List shapefile attachments for rows in range."""
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    try:
        r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(),
            params={"include": "attachments"},
            timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    def cell_val_by_id(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    # EN: If frontend sends explicit row numbers (filtered view), use those directly
    # ES: Si el frontend envía números de fila explícitos (vista filtrada), usarlos directamente
    row_numbers = body.get("rowNumbers")
    if row_numbers:
        rn_set = set(row_numbers)
        target_rows = [
            row for row in sheet.get("rows", [])
            if row.get("rowNumber", 0) in rn_set
        ]
    else:
        target_rows = [
            row for row in sheet.get("rows", [])
            if row_start <= row.get("rowNumber", 0) <= row_end
        ]
        if comp == "C2":
            target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val_by_id)

    results = []
    for row in target_rows:
        rn = row.get("rowNumber", 0)
        atts = row.get("attachments", [])
        if atts:
            results.append({
                "rowNumber": rn,
                "attachments": [{"id": a["id"], "name": a.get("name", "")} for a in atts],
            })
    return jsonify(rows=results, sheetId=sid)


@app.route("/api/smartsheet/download-attachment", methods=["POST"])
def ss_download_attachment():
    """Download a single attachment to local disk."""
    body = request.json or {}
    sid = body.get("sheetId")
    att_id = body.get("attachmentId")
    if not sid or not att_id:
        return jsonify(error="sheetId and attachmentId required"), 400

    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    os.makedirs(dest_folder, exist_ok=True)

    # Get direct download URL
    r = smartsheet_request(
        "GET",
        f"{SS_BASE}/sheets/{sid}/attachments/{att_id}",
        headers=_ss_headers(),
        timeout=30,
    )
    if not r.ok:
        return jsonify(error=f"Failed to get attachment info: {r.status_code}"), r.status_code
    att_info = r.json()
    url = att_info.get("url")
    name = att_info.get("name", f"attachment_{att_id}")

    # Download file
    dl = requests.get(url, timeout=60)
    if not dl.ok:
        return jsonify(error=f"Download failed: {dl.status_code}"), 500

    safe_name = re.sub(r'[^\w.\-]', '_', name)
    filepath = os.path.join(dest_folder, safe_name)
    with open(filepath, "wb") as f:
        f.write(dl.content)

    return jsonify(downloaded=safe_name, path=filepath, size=len(dl.content))


# ---------------------------------------------------------------------------
# API: ArcPy Script Generator  (PASO 3 + PASO 4 + PASO 6)
# ---------------------------------------------------------------------------

@app.route("/api/arcpy/generate-script", methods=["POST"])
def arcpy_generate_script():
    """Generate a ready-to-run ArcPy script with user parameters."""
    body = request.json or {}
    step = body.get("step", "")
    params = body.get("params", {})

    # Inject .env defaults for AGOL if not provided by the client
    params.setdefault("agol_item_id", os.getenv("AGOL_POLYGON_ITEM_ID", ""))
    params.setdefault("agol_polygon_item_id", os.getenv("AGOL_POLYGON_ITEM_ID", ""))
    params.setdefault("agol_point_item_id", os.getenv("AGOL_POINT_ITEM_ID", ""))
    params.setdefault("agol_portal_url", os.getenv("ARCGIS_ORG_URL", ""))

    # Unified registry: PASO 3 + PASO 4 + PASO 6 scripts + legacy import
    all_scripts = {}
    all_scripts.update(PASO3_SCRIPTS)
    all_scripts.update(PASO4_SCRIPTS)
    all_scripts.update(PASO6_SCRIPTS)
    all_scripts["import"] = _script_import_to_arcgis

    gen = all_scripts.get(step)
    if gen is None:
        return jsonify(error=f"Unknown step: {step}. Available: {sorted(all_scripts.keys())}"), 400
    return jsonify(script=gen(params), step=step)


# ---------------------------------------------------------------------------
# API: 3-Way Verification  (PASO 4 — G6)
# ---------------------------------------------------------------------------

@app.route("/api/verify/3way", methods=["POST"])
def verify_3way():
    """
    Server-side Smartsheet vs GIS comparison (AGOL part runs via arcpy script).

    Body JSON:
      - ss_data:    list of Smartsheet row dicts (CÓDIGO_DE_LA_ACTIVIDAD, TOTAL_DE_HECTÁREAS)
      - component:  "C1" | "C2" | "C3"  (optional — for Smartsheet fetch if ss_data absent)
      - rowStart, rowEnd: row range for SS fetch
      - gis_records: list of {CdgActvdd, Area_ha} dicts from a GIS CSV (optional)

    Returns:
      { ss: {count, area_total, codes: [...]},
        gis: {count, area_total, codes: [...]},
        ss_only: [...], gis_only: [...],
        area_delta: float, status: "ok"|"warning"|"error" }
    """
    body = request.json or {}

    # --- Build SS dataset (from body or live fetch) ---
    ss_data = body.get("ss_data")
    if not ss_data:
        comp = body.get("component", "C1")
        row_start = body.get("rowStart", 1)
        row_end = body.get("rowEnd", 200)
        sid = _sheet_id(comp)
        if not sid:
            return jsonify(friendly_error("no_sheet_id")), 400
        try:
            r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
        except requests.ConnectionError:
            return jsonify(friendly_error("ss_network")), 503
        if r.status_code == 401:
            return jsonify(friendly_error("ss_401")), 401
        if not r.ok:
            return jsonify(friendly_error("unknown")), r.status_code
        sheet = r.json()
        col_map = {c["id"]: c["title"] for c in sheet.get("columns", [])}
        ss_data = []
        for row in sheet.get("rows", []):
            rn = row.get("rowNumber", 0)
            if rn < row_start or rn > row_end:
                continue
            cells = {}
            for cell in row.get("cells", []):
                title = col_map.get(cell.get("columnId"))
                if title:
                    cells[title] = cell.get("value")
            ss_data.append(cells)

    # --- Parse SS records ---
    ss_records = {}
    for row in (ss_data or []):
        code = None
        area = 0.0
        for k, v in row.items():
            if "DIGO" in k.upper() or "CODIGO" in k.upper():
                code = str(v).strip() if v else None
            elif "HECT" in k.upper() or "AREA" in k.upper():
                try:
                    area = float(str(v).replace(",", ".")) if v else 0.0
                except ValueError:
                    area = 0.0
        if code:
            ss_records[code] = ss_records.get(code, 0.0) + area

    # --- Parse GIS records (from body if provided) ---
    gis_records = {}
    for rec in (body.get("gis_records") or []):
        code = str(rec.get("CdgActvdd") or rec.get("codigo") or "").strip()
        try:
            area = float(rec.get("Area_ha") or rec.get("area_ha") or 0)
        except (ValueError, TypeError):
            area = 0.0
        if code:
            gis_records[code] = gis_records.get(code, 0.0) + area

    # --- Compute comparison ---
    ss_set  = set(ss_records.keys())
    gis_set = set(gis_records.keys())
    ss_only  = sorted(ss_set  - gis_set)
    gis_only = sorted(gis_set - ss_set)

    ss_area_total  = sum(ss_records.values())
    gis_area_total = sum(gis_records.values())
    area_delta = abs(gis_area_total - ss_area_total)

    status = "ok"
    if ss_only:
        status = "error"
    elif gis_only or area_delta > 0.1:
        status = "warning"

    return jsonify(
        ss=dict(count=len(ss_set),  area_total=round(ss_area_total, 2),  codes=sorted(ss_set)),
        gis=dict(count=len(gis_set), area_total=round(gis_area_total, 2), codes=sorted(gis_set)),
        ss_only=ss_only,
        gis_only=gis_only,
        area_delta=round(area_delta, 2),
        status=status,
    )


@app.route("/api/verify/3way/script", methods=["POST"])
def verify_3way_script():
    """
    Generate a full 3-way arcpy+arcgis verification script for ArcGIS Pro.

    Body JSON: same params as script_verify_3way_sync()
    Returns: { script: "..." }
    """
    from paso4_scripts import script_verify_3way_sync
    body = request.json or {}
    # Inject .env defaults for AGOL if not provided by the client
    cfg = _cfg()
    body.setdefault("agol_item_id", os.getenv("AGOL_POLYGON_ITEM_ID", ""))
    body.setdefault("agol_portal_url", os.getenv("ARCGIS_ORG_URL", ""))
    if not body.get("agol_url"):
        body.setdefault("agol_url", cfg.get("AGOL_POLYGON_URL", ""))
    return jsonify(script=script_verify_3way_sync(body))


def _script_import_to_arcgis(p):
    gdb = p.get("gdb", r"C:\Path\To\Your.gdb")
    csv_path = p.get("csvPath", r"C:\Path\To\export.csv")
    comp = p.get("component", "C1")
    ts = datetime.now().strftime("%Y%m%d%H%M")
    table_name = f"ss_{comp}_{ts}"
    return textwrap.dedent(f"""\
        # Import Smartsheet CSV to ArcGIS Geodatabase
        import arcpy

        csv_path = r"{csv_path}"
        gdb = r"{gdb}"
        table_name = "{table_name}"

        arcpy.env.workspace = gdb
        arcpy.conversion.ExportTable(csv_path, table_name)
        print(f"Imported {{table_name}} into {{gdb}}")
    """)


# ---------------------------------------------------------------------------
# API: Power BI Pipeline  (PASO 6)
# ---------------------------------------------------------------------------

@app.route("/api/pbi/status")
def pbi_status():
    """Return Power BI Desktop status: running, port, basepath, paths."""
    from pbi_refresh import find_pbi_port, is_port_open, PBIP_PATH, TMDL_FOLDER

    port = find_pbi_port(pbip_path=PBIP_PATH)
    running = bool(port and is_port_open(port))

    # Read BasePath from expressions.tmdl
    basepath = None
    basepath_exists = False
    expr_file = os.path.join(TMDL_FOLDER, "expressions.tmdl")
    if os.path.exists(expr_file):
        with open(expr_file, "r", encoding="utf-8") as f:
            for line in f:
                if "BasePath" in line and '"' in line:
                    start = line.find('"')
                    end = line.find('"', start + 1)
                    if start > 0 and end > start:
                        basepath = line[start + 1:end]
                        basepath_exists = os.path.exists(basepath)
                    break

    return jsonify(
        running=running,
        port=port,
        connectionString=f"Provider=MSOLAP;Data Source=localhost:{port}" if running else None,
        pbipPath=PBIP_PATH,
        pbipExists=os.path.exists(PBIP_PATH),
        tmdlExists=os.path.exists(os.path.join(TMDL_FOLDER, "database.tmdl")),
        basePath=basepath,
        basePathExists=basepath_exists,
    )


@app.route("/api/pbi/launch", methods=["POST"])
def pbi_launch():
    """Launch Power BI Desktop with the .pbip file."""
    from pbi_refresh import launch_pbi_desktop, find_pbi_executable, PBIP_PATH

    exe = find_pbi_executable()
    if not exe:
        return jsonify(friendly_error("powerbi_no_license")), 400

    if not os.path.exists(PBIP_PATH):
        # ES: Archivo .pbip no encontrado / EN: .pbip file not found
        return jsonify(error="PBIP file not found", path=PBIP_PATH), 404

    try:
        subprocess.Popen([exe, PBIP_PATH])
        return jsonify(ok=True, message="Power BI Desktop launched / PBI Desktop abierto.")
    except Exception as exc:
        return jsonify(error=str(exc)), 500


@app.route("/api/pbi/wait", methods=["GET", "POST"])
def pbi_wait():
    """Poll for PBI Desktop readiness (one shot, no long-polling)."""
    from pbi_refresh import find_pbi_port, is_port_open, PBIP_PATH

    port = find_pbi_port(pbip_path=PBIP_PATH)
    if port and is_port_open(port):
        return jsonify(
            ready=True,
            port=port,
            connectionString=f"Provider=MSOLAP;Data Source=localhost:{port}",
        )
    return jsonify(ready=False, port=None)


@app.route("/api/pbi/tables")
def pbi_tables():
    """Return the list of refreshable tables from the pipeline doc."""
    tables = {
        "excel": [
            {"table": "Area_Indicator_ID", "source": "Area_Indicator_ID.xlsx"},
            {"table": "BASE_Indicator", "source": "Tbl_Integrado.xlsx"},
            {"table": "Index_Pp", "source": "Tbl_Integrado.xlsx"},
            {"table": "KOICA_ORG_DIM_TARGET", "source": "KOICA_DATA_MODEL_v1.xlsx"},
            {"table": "KOICA_ORG_FACT_AREA", "source": "KOICA_DATA_MODEL_v1.xlsx"},
            {"table": "KOICA_ORG_FACT_BNF", "source": "KOICA_DATA_MODEL_v1.xlsx"},
            {"table": "KOICA_ORG_FACT_ORG", "source": "KOICA_DATA_MODEL_v1.xlsx"},
            {"table": "KOICA_ORG_FACT_PRJ_OBJ", "source": "KOICA_DATA_MODEL_v1.xlsx"},
            {"table": "Location_Grants", "source": "Grants_X_YDD.xls"},
            {"table": "Tbl_Actvdd_m_e_data", "source": "Tbl_Integrado.xlsx"},
            {"table": "Tbl_Actvdd_TIPO_me", "source": "Tbl_Integrado.xlsx"},
            {"table": "Tbl_Gender", "source": "data_structure_Gender.xlsx"},
            {"table": "Tbl_Pp", "source": "Tbl_Integrado.xlsx"},
            {"table": "Tbl_Grants", "source": "Tbl_Integrado.xlsx"},
        ],
        "folder": [
            {"table": "Tbl_Area", "source": "AR_3_Area/0. Database_oficial"},
            {"table": "Tbl_Area_poli", "source": "AR_3_Area/0. Database_oficial"},
            {"table": "Comités-Microcuencas", "source": "Monster-File/.../Comités-Microcuencas"},
            {"table": "ESMP_1_2", "source": "Monster-File/.../1.2. Comités-Microcuencas"},
            {"table": "ESMP_G_1", "source": "Monster-File/.../1.2. Comités-Microcuencas"},
        ],
        "smartsheet": [
            {"table": "Tbl_Actvdd_ss", "source": "3 sheets combinados (7709708342585220)"},
        ],
        "erosion": [
            {"table": "ZonalSt_SWY_BL_aet", "source": "SWY_indicadores_BalanceH_LB.xlsx"},
            {"table": "ZonalSt_SWY_BL_L", "source": "SWY_indicadores_BalanceH_LB.xlsx"},
            {"table": "ZonalSt_SWY_BL_P", "source": "SWY_indicadores_BalanceH_LB.xlsx"},
            {"table": "ZonalSt_SWY_BL_QF", "source": "SWY_indicadores_BalanceH_LB.xlsx"},
        ],
    }
    return jsonify(tables)


# ---------------------------------------------------------------------------
# API: Batch Download  (PASO 1 completion)
# ---------------------------------------------------------------------------

@app.route("/api/quarterly/create", methods=["POST"])
def quarterly_create():
    """Pre-create quarterly folder structure without downloading.
    / Pre-crea estructura de carpetas trimestral sin descargar.

    Body: { "baseDir": str, "quarter": str, "component": str }
    """
    from paso1_quarterly import create_quarterly_structure

    body = request.json or {}
    quarter = body.get("quarter", "")
    component = body.get("component", "C1")

    if not quarter:
        return jsonify(error="quarter is required / el trimestre es obligatorio"), 400

    try:
        base_dir = str(safe_resolve(
            body.get("baseDir", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    try:
        folder = create_quarterly_structure(base_dir, quarter, component)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400

    return jsonify(ok=True, folder=folder)


@app.route("/api/smartsheet/batch-download", methods=["POST"])
def ss_batch_download():
    """Download all shapefile ZIPs with proper naming per component.

    Supports checkpoint/resume: saves progress to a JSON file so that
    interrupted downloads can be resumed from where they stopped.
    """
    import json as _json
    from ar_utils import extract_abe_hint_from_filename, extract_hectares_from_filename, extract_and_validate_zip, is_c1_row_shape_zip, is_shape_zip_attachment, validate_abe
    from paso1_quarterly import place_shapefile

    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    quarter = body.get("quarter", "").strip()
    resume = body.get("resume", False)

    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    os.makedirs(dest_folder, exist_ok=True)

    # -- Checkpoint file lives in dest_folder
    ckpt_path = os.path.join(dest_folder, ".batch_checkpoint.json")

    def _load_checkpoint():
        if resume and os.path.isfile(ckpt_path):
            try:
                with open(ckpt_path, "r", encoding="utf-8") as f:
                    return _json.load(f)
            except Exception:
                pass
        return {"completed": {}, "results": []}

    def _save_checkpoint(ckpt):
        try:
            with open(ckpt_path, "w", encoding="utf-8") as f:
                _json.dump(ckpt, f, ensure_ascii=False, indent=1)
        except Exception:
            pass

    ckpt = _load_checkpoint()
    completed_keys = set(ckpt.get("completed", {}).keys())

    try:
        r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(),
            params={"include": "attachments"},
            timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    def cell_val(row, title):
        cid = col_map.get(title)
        if not cid:
            return None
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    def cell_val_by_id(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    row_numbers = body.get("rowNumbers")
    if row_numbers:
        rn_set = set(row_numbers)
        target_rows = [
            row for row in sheet.get("rows", [])
            if row.get("rowNumber", 0) in rn_set
        ]
    else:
        target_rows = [
            row for row in sheet.get("rows", [])
            if row_start <= row.get("rowNumber", 0) <= row_end
        ]
        if comp == "C2":
            target_rows = _c2_filter_rows(target_rows, body, col_map, cell_val_by_id)

    # -- Build task list & validate ABE values / Construir tareas y validar AbE
    tasks = []
    abe_warnings = []  # rows with invalid ABE values / filas con valores AbE inválidos

    # For C2: build a map from summary row → child rows (same contract+org+quarter)
    # / Para C2: construir mapa de fila resumen → filas hijas (mismo contrato+org+trimestre)
    all_rows_by_rn = {r.get("rowNumber", 0): r for r in sheet.get("rows", [])}

    def _c2_child_abe_values(summary_row):
        """Collect AbE values from child rows following a C2 summary row.
        Uses fill-down semantics: child rows with empty CONTRATO/ORG/QT
        inherit the parent value. Stops when a *different* non-empty value
        appears.
        / Usa semántica fill-down: filas hijas con CONTRATO/ORG/QT vacíos
        heredan del padre. Se detiene cuando aparece un valor diferente no vacío.
        """
        contract = cell_val(summary_row, "NÚMERO DE CONTRATO") or ""
        org = cell_val(summary_row, "ORGANIZACIÓN") or ""
        qt = cell_val(summary_row, "TRIMESTRE QUE REPORTA") or ""
        start_rn = summary_row.get("rowNumber", 0)
        abe_vals = []
        for rn in range(start_rn + 1, start_rn + 50):  # scan next rows
            child = all_rows_by_rn.get(rn)
            if not child:
                break
            c_contract = cell_val(child, "NÚMERO DE CONTRATO") or ""
            c_org = cell_val(child, "ORGANIZACIÓN") or ""
            c_qt = cell_val(child, "TRIMESTRE QUE REPORTA") or ""
            # Stop when a different non-empty value appears (new group)
            # Empty values are inherited from the summary row (fill-down)
            # / Detenerse cuando aparece un valor diferente no vacío (nuevo grupo).
            #   Los valores vacíos se heredan de la fila resumen (fill-down).
            if c_contract and c_contract != contract:
                break
            if c_org and c_org != org:
                break
            if c_qt and c_qt != qt:
                break
            child_abe = cell_val(child, "ACCIONES DE RESTAURACIÓN AbE") or ""
            child_code = cell_val(child, "CÓDIGO DE LA ACTIVIDAD") or ""
            child_ha = cell_val(child, "TOTAL DE HECTÁREAS")
            child_fecha = cell_val(child, "FECHA DE LA ACTIVIDAD") or ""
            child_nombre = cell_val(child, "NOMBRE DE LA ACTIVIDAD") or ""
            if child_code or child_abe:
                child_rn = child.get("rowNumber", rn)
                abe_vals.append({
                    "row": child_rn,
                    "value": child_abe,
                    "code": child_code,
                    "hectares": float(child_ha) if child_ha else 0.0,
                    "fecha": child_fecha,
                    "nombre": child_nombre,
                })
        return abe_vals

    # Track seen attachment names per identifier to prevent the same ZIP
    # from being processed twice (e.g. when both a C2 summary row and its
    # child row carry the same attachment).
    # / Rastrear nombres de adjuntos vistos por identificador para evitar
    #   que el mismo ZIP se procese dos veces.
    _seen_att = set()  # (identifier, att_name)

    for row in target_rows:
        rn = row.get("rowNumber", 0)
        # Row-level context used by the C1/C3 detector to accept a generic
        # ZIP name when it is the only ZIP and the row carries hectares + AbE.
        # / Contexto a nivel de fila usado por el detector C1/C3: aceptar un
        # ZIP con nombre genérico si es el único y la fila tiene hectáreas + AbE.
        row_zip_count = sum(
            1 for _a in row.get("attachments", [])
            if (_a.get("name", "") or "").lower().endswith(".zip")
        )
        row_has_hectares = cell_val(row, "TOTAL DE HECTÁREAS") not in (None, "")
        row_has_abe = bool(cell_val(row, "ACCIONES DE RESTAURACIÓN AbE"))
        for att in row.get("attachments", []):
            att_name = att.get("name", "")
            if comp == "C2":
                # C2 keeps the existing filename-only rule because a summary
                # row often carries multiple SHP ZIPs differentiated by name.
                if not is_shape_zip_attachment(att_name):
                    continue
            else:
                if not is_c1_row_shape_zip(
                    att_name, row_zip_count, row_has_hectares, row_has_abe,
                ):
                    continue
            if comp == "C2":
                contract = cell_val(row, "NÚMERO DE CONTRATO") or ""
                org = cell_val(row, "ORGANIZACIÓN") or ""
                row_quarter = cell_val(row, "TRIMESTRE QUE REPORTA") or ""
                # Fill-down: if child row has empty group fields, scan backwards
                # to find the nearest summary row values.
                # / Fill-down: si la fila hija tiene campos vacíos, buscar hacia
                #   atrás la fila resumen más cercana.
                if not contract or not org or not row_quarter:
                    for prev_rn in range(rn - 1, max(rn - 50, 0), -1):
                        prev = all_rows_by_rn.get(prev_rn)
                        if not prev:
                            continue
                        pc = cell_val(prev, "NÚMERO DE CONTRATO") or ""
                        po = cell_val(prev, "ORGANIZACIÓN") or ""
                        pq = cell_val(prev, "TRIMESTRE QUE REPORTA") or ""
                        if pc and not contract:
                            contract = pc
                        if po and not org:
                            org = po
                        if pq and not row_quarter:
                            row_quarter = pq
                        if contract and org and row_quarter:
                            break
                base_name = f"{contract}_{org}_{row_quarter}".replace(" ", "_")
            else:
                base_name = cell_val(row, "CÓDIGO DE LA ACTIVIDAD") or f"row_{rn}"

            if comp == "C2":
                # If this row itself has TOTAL DE HECTÁREAS and a CODE, it's a
                # child row with SHP attached — use its own values directly.
                # / Si esta fila tiene hectáreas y código, es una fila hija con
                #   SHP adjunto — usar sus propios valores directamente.
                own_ha = cell_val(row, "TOTAL DE HECTÁREAS")
                own_code = cell_val(row, "CÓDIGO DE LA ACTIVIDAD") or ""
                own_abe = cell_val(row, "ACCIONES DE RESTAURACIÓN AbE") or ""
                if own_ha is not None and own_code:
                    abe_val = own_abe
                    c2_children = [{
                        "row": rn,
                        "value": own_abe,
                        "code": own_code,
                        "hectares": float(own_ha) if own_ha else 0.0,
                        "fecha": cell_val(row, "FECHA DE LA ACTIVIDAD") or "",
                        "nombre": cell_val(row, "NOMBRE DE LA ACTIVIDAD") or "",
                    }]
                else:
                    # Summary row path: collect child AbE values
                    # / Ruta de fila resumen: recopilar valores AbE de hijas
                    abe_val = own_abe
                    child_abe_list = _c2_child_abe_values(row)
                    c2_children = child_abe_list
                    if child_abe_list:
                        # Validate each child's AbE
                        for ca in child_abe_list:
                            if ca.get("value"):
                                abe_check = validate_abe(ca["value"])
                                if not abe_check["valid"]:
                                    abe_warnings.append({
                                        "row": ca["row"],
                                        "identifier": base_name,
                                        "abe_value": ca["value"],
                                        "message": abe_check["message"],
                                    })
                        # Use first child AbE that has a value for naming
                        # / Usar el primer AbE de fila hija que tenga valor
                        if not abe_val:
                            for ca in child_abe_list:
                                if ca.get("value"):
                                    abe_val = ca["value"]
                                    break
                    else:
                        # Validate the summary row's own AbE value
                        abe_check = validate_abe(abe_val)
                        if not abe_check["valid"]:
                            abe_warnings.append({
                                "row": rn,
                                "identifier": base_name,
                                "abe_value": abe_val,
                                "message": abe_check["message"],
                            })
            else:
                c2_children = []
                abe_val = cell_val(row, "ACCIONES DE RESTAURACIÓN AbE") or ""
                # Validate ABE / Validar AbE
                abe_check = validate_abe(abe_val)
                if not abe_check["valid"]:
                    abe_warnings.append({
                        "row": rn,
                        "identifier": base_name,
                        "abe_value": abe_val,
                        "message": abe_check["message"],
                    })

            safe_name = re.sub(r'[^\w.\-]', '_', base_name) + ".zip"

            # -- C2 multi-SHP: extract ABE hint from original attachment name
            # to differentiate files and pass correct abe_value to place_shapefile.
            # / C2 multi-SHP: extraer pista AbE del nombre original para diferenciar.
            orig_att_name = att.get("name", "")
            abe_hint = ""
            ha_hint = None
            if comp == "C2":
                abe_hint = extract_abe_hint_from_filename(orig_att_name)
                ha_hint = extract_hectares_from_filename(orig_att_name)
                if abe_hint:
                    # Differentiate safe_name so multiple SHP ZIPs from the same
                    # row don't overwrite each other.
                    safe_name = re.sub(r'[^\w.\-]', '_', base_name) + f"-{abe_hint}.zip"
                    # Find the matching child ABE full value for this hint
                    # so place_shapefile uses the correct ABE abbreviation.
                    # / Buscar el valor AbE completo del hijo que coincida con la pista.
                    if c2_children:
                        from ar_utils import abe as _abe_fn
                        if abe_hint == "saf":
                            # SAF is a generic prefix for all "Sistema
                            # Agroforestal" subtypes.  Use ha_hint to pick
                            # the best matching child by hectares.
                            best_ca, best_r = None, 0.0
                            for ca in c2_children:
                                v = (ca.get("value") or "").lower()
                                if "sistema agroforestal" not in v:
                                    continue
                                ca_ha = float(ca.get("hectares", 0) or 0)
                                if ca_ha > 0 and ha_hint and ha_hint > 0:
                                    r = min(ha_hint, ca_ha) / max(ha_hint, ca_ha)
                                    if r > best_r:
                                        best_r, best_ca = r, ca
                                elif best_ca is None:
                                    best_ca = ca
                            if best_ca:
                                abe_val = best_ca["value"]
                        else:
                            for ca in c2_children:
                                if ca.get("value") and _abe_fn(ca["value"]) == abe_hint:
                                    abe_val = ca["value"]
                                    break

            # Deduplicate: skip if same attachment name already queued for
            # this identifier (summary row + child row both having same file)
            _att_key = (base_name, att_name)
            if _att_key in _seen_att:
                continue
            _seen_att.add(_att_key)

            tasks.append({
                "row": rn, "att_id": att["id"],
                "safe_name": safe_name,
                "identifier": base_name,  # single folder per C2 row
                "abe_value": abe_val,
                "abe_hint": abe_hint,
                "ha_hint": ha_hint,
                "orig_att_name": orig_att_name,
                "c2_children": c2_children,
            })

    total = len(tasks)
    results = list(ckpt.get("results", []))  # carry over previous results on resume
    skipped = 0

    for idx, task in enumerate(tasks):
        rn = task["row"]
        safe_name = task["safe_name"]
        identifier = task["identifier"]
        abe_val = task.get("abe_value", "")
        ckpt_key = f"{rn}:{task['att_id']}"

        # Skip already-completed items on resume
        if ckpt_key in completed_keys:
            skipped += 1
            continue

        # Download attachment metadata
        att_r = smartsheet_request(
            "GET",
            f"{SS_BASE}/sheets/{sid}/attachments/{task['att_id']}",
            headers=_ss_headers(), timeout=30,
        )
        if not att_r.ok:
            entry = {"name": safe_name, "row": rn, "ok": False,
                     "error": f"Meta fetch failed: {att_r.status_code}"}
            results.append(entry)
            ckpt["results"] = results
            ckpt["last_index"] = idx
            ckpt["total"] = total
            _save_checkpoint(ckpt)
            continue

        url = att_r.json().get("url")
        zip_path = os.path.join(dest_folder, safe_name)
        try:
            stream_download_to_file(url, zip_path, timeout=60)
        except requests.HTTPError as dl_err:
            entry = {"name": safe_name, "row": rn, "ok": False,
                     "error": f"Download failed: {dl_err.response.status_code if dl_err.response else 'unknown'}"}
            results.append(entry)
            ckpt["results"] = results
            ckpt["last_index"] = idx
            ckpt["total"] = total
            _save_checkpoint(ckpt)
            continue
        except (requests.ConnectionError, requests.Timeout) as dl_err:
            entry = {"name": safe_name, "row": rn, "ok": False,
                     "error": f"Download error: {dl_err}"}
            results.append(entry)
            ckpt["results"] = results
            ckpt["last_index"] = idx
            ckpt["total"] = total
            _save_checkpoint(ckpt)
            continue

        if quarter:
            placement = place_shapefile(zip_path, dest_folder, identifier, comp,
                                       abe_value=abe_val,
                                       orig_att_name=task.get("orig_att_name", ""))
            entry = {
                "name": safe_name, "row": rn,
                "ok": placement["ok"],
                "shapefiles": [fn for fn in placement["files"] if fn.lower().endswith(".shp")],
                "errors": placement["errors"],
                "folder": placement["folder"],
            }
            # C2: save child-row mapping for CdgActvdd assignment in ArcPy
            # / C2: guardar mapeo de filas hijas para asignación de CdgActvdd en ArcPy
            c2_children = task.get("c2_children", [])
            if comp == "C2" and c2_children and placement["ok"]:
                # Always write the FULL child list — the ArcPy 3-tier matching
                # (_score_mapping) filters by ABE as its first step.  Since
                # multiple SHP ZIPs share one folder, only one mapping file
                # is needed with all children.
                # / Siempre escribir la lista COMPLETA de hijos — el matching
                # ArcPy filtra por AbE como primer paso.
                import json as _json2
                mapping_path = os.path.join(placement["folder"], "_cdg_mapping.json")
                # Include metadata: summary-row AbE and sheet permalink
                # so the ArcPy script can use summary AbE as fallback
                # and log Smartsheet links for review.
                mapping_data = {
                    "_meta": {
                        "summary_abe": task.get("abe_value", ""),
                        "sheet_id": str(sid),
                        "permalink": sheet.get("permalink", ""),
                        "summary_row": rn,
                        "identifier": identifier,
                        "ha_hint": task.get("ha_hint"),
                    },
                    "children": c2_children,
                }
                try:
                    with open(mapping_path, "w", encoding="utf-8") as mf:
                        _json2.dump(mapping_data, mf, ensure_ascii=False, indent=2)
                except OSError:
                    pass
        else:
            extract_dir = os.path.join(dest_folder, os.path.splitext(safe_name)[0])
            validation = extract_and_validate_zip(zip_path, extract_dir)
            entry = {
                "name": safe_name, "row": rn,
                "ok": validation["ok"],
                "shapefiles": [os.path.basename(s) for s in validation["shapefiles"]],
                "errors": validation["errors"],
                "folder": os.path.abspath(extract_dir),
            }

        # Remove original ZIP after successful extraction
        # / Eliminar ZIP original después de extracción exitosa
        if entry["ok"]:
            try:
                os.remove(zip_path)
            except OSError:
                pass

        results.append(entry)
        ckpt["completed"][ckpt_key] = True
        ckpt["results"] = results
        ckpt["last_index"] = idx
        ckpt["total"] = total
        _save_checkpoint(ckpt)

    ok_count = sum(1 for r in results if r["ok"])
    all_done = (skipped + len(results) - len(ckpt.get("results", [])) + skipped) >= 0

    # Clean up checkpoint when everything succeeded
    if ok_count == len(results) and len(results) == total:
        try:
            os.remove(ckpt_path)
        except OSError:
            pass

    return jsonify(
        downloaded=len(results), valid=ok_count, total=total,
        skipped=skipped, resumed=resume and skipped > 0,
        results=results, dest_folder=dest_folder,
        abe_warnings=abe_warnings,
    )


@app.route("/api/smartsheet/excel-area-download", methods=["POST"])
def ss_excel_area_download():
    """Download *area* Excel attachments for rows missing a shapefile ZIP.

    / Descarga adjuntos Excel de áreas para filas sin shapefile ZIP.

    Used as a fallback by the Paso 1 agent: when ``/api/smartsheet/batch-download``
    returns no shapefile attachments for a row, the agent calls this endpoint
    with the same ``rowNumbers``.  Each matching ``.xls/.xlsx`` whose name
    contains area/hectare keywords is downloaded into
    ``{destFolder}/{base_name}/`` and returned for downstream Excel-to-point
    conversion.

    Body: ``{ component, rowNumbers, destFolder, quarter }``
    """
    from ar_utils import is_area_excel_attachment

    body = request.json or {}
    comp = body.get("component", "C1")
    quarter = (body.get("quarter") or "").strip()

    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    os.makedirs(dest_folder, exist_ok=True)

    try:
        r = smartsheet_request(
            "GET", f"{SS_BASE}/sheets/{sid}",
            headers=_ss_headers(), params={"include": "attachments"}, timeout=30,
        )
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["title"]: c["id"] for c in sheet.get("columns", [])}

    def cell_val(row, title):
        cid = col_map.get(title)
        if not cid:
            return None
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    row_numbers = body.get("rowNumbers") or []
    rn_set = set(row_numbers)
    target_rows = [
        row for row in sheet.get("rows", [])
        if row.get("rowNumber", 0) in rn_set
    ]

    results = []
    for row in target_rows:
        rn = row.get("rowNumber", 0)
        # base_name follows the same convention used by batch-download so
        # downstream tools can find the folder identically.
        # / base_name sigue la misma convención que batch-download.
        if comp == "C2":
            contract = cell_val(row, "NÚMERO DE CONTRATO") or ""
            org = cell_val(row, "ORGANIZACIÓN") or ""
            row_quarter = cell_val(row, "TRIMESTRE QUE REPORTA") or ""
            base_name = f"{contract}_{org}_{row_quarter}".replace(" ", "_")
        else:
            base_name = cell_val(row, "CÓDIGO DE LA ACTIVIDAD") or f"row_{rn}"

        safe_id = re.sub(r"[^\w.\-]", "_", str(base_name).strip())
        row_folder = os.path.join(dest_folder, safe_id)
        os.makedirs(row_folder, exist_ok=True)

        excel_attachments = [
            att for att in row.get("attachments", [])
            if is_area_excel_attachment(att.get("name", ""))
        ]
        if not excel_attachments:
            results.append({
                "row": rn, "ok": False, "code": base_name,
                "folder": os.path.abspath(row_folder),
                "files": [], "error": "no_excel_attachment",
            })
            continue

        downloaded: list[str] = []
        download_err = ""
        for att in excel_attachments:
            att_r = smartsheet_request(
                "GET",
                f"{SS_BASE}/sheets/{sid}/attachments/{att['id']}",
                headers=_ss_headers(), timeout=30,
            )
            if not att_r.ok:
                download_err = f"meta_{att_r.status_code}"
                continue
            url = att_r.json().get("url")
            orig_name = att.get("name", "area.xlsx")
            safe_name = re.sub(r"[^\w.\-]", "_", orig_name)
            local_path = os.path.join(row_folder, safe_name)
            try:
                stream_download_to_file(url, local_path, timeout=60)
                downloaded.append(os.path.abspath(local_path))
            except (requests.HTTPError, requests.ConnectionError,
                    requests.Timeout) as dl_err:
                download_err = f"dl_{dl_err}"

        results.append({
            "row": rn,
            "ok": bool(downloaded),
            "code": base_name,
            "folder": os.path.abspath(row_folder),
            "files": downloaded,
            "error": download_err if not downloaded else "",
        })

    return jsonify(results=results, dest_folder=dest_folder)


@app.route("/api/arcpy/run-excel-to-point", methods=["POST"])
def arcpy_run_excel_to_point():
    """Convert area Excel(s) to point feature classes and add to ArcGIS Pro.

    / Convierte Excel(s) de áreas en feature classes de puntos y los agrega
    al mapa de ArcGIS Pro.

    Body: ``{ destFolder, quarter, items: [{folder, cdg}], lang }``
    Each item points to a folder containing an area Excel; the endpoint
    parses each Excel into a CSV, then runs an ArcPy script that creates
    a point FC in ``CSVPOINT_TO_GDB`` and adds it to the configured map.
    """
    import json as _json
    import tempfile
    from paso1_excel_to_point import (
        find_area_excel, parse_area_excel, script_excel_to_point,
    )

    arcpy_python = os.getenv("ARCPY_PYTHON", "")
    if not arcpy_python or not os.path.isfile(arcpy_python):
        return jsonify(friendly_error("arcpy_not_configured")), 400

    body = request.json or {}
    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    quarter = (body.get("quarter") or "").strip()
    items = body.get("items") or []
    aprx = os.getenv("APRX", "")
    map_name = os.getenv("AR_MAP_NAME", "AR_EbA_Area")
    gdb_path = os.getenv("CSVPOINT_TO_GDB", "")

    if not aprx:
        return jsonify(error="APRX not configured in .env"), 400
    if not gdb_path:
        return jsonify(error="CSVPOINT_TO_GDB not configured in .env"), 400

    # Resolve every item folder and parse its Excel into a CSV.
    # / Resolver carpeta de cada ítem y parsear su Excel a CSV.
    csv_tasks = []
    parse_errors = []
    for it in items:
        folder = it.get("folder") or ""
        cdg = it.get("cdg") or ""
        if not folder or not cdg:
            parse_errors.append({"cdg": cdg, "error": "missing_folder_or_cdg"})
            continue
        try:
            safe_folder = str(safe_resolve(folder, _allowed_roots()))
        except ValueError as e:
            parse_errors.append({"cdg": cdg, "error": f"path_unsafe: {e}"})
            continue
        excel_path = find_area_excel(safe_folder)
        if not excel_path:
            parse_errors.append({"cdg": cdg, "error": "no_excel_in_folder",
                                 "folder": safe_folder})
            continue
        parsed = parse_area_excel(excel_path, cdg)
        if not parsed["ok"]:
            parse_errors.append({"cdg": cdg, "error": "parse_failed",
                                 "details": parsed["errors"]})
            continue
        csv_tasks.append({
            "csv_path": parsed["csv_path"], "cdg": cdg, "quarter": quarter,
        })

    if not csv_tasks:
        return jsonify(
            ok=False, error="no_valid_excel", parseErrors=parse_errors,
        ), 400

    script = script_excel_to_point({
        "csvs": csv_tasks, "gdb_path": gdb_path,
        "aprx": aprx, "map_name": map_name,
        "lang": body.get("lang", "es"),
    })

    script_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".py", prefix="ar_excel_to_point_",
            delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write(script)
            script_path = tmp.name

        env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
        result = subprocess.run(
            [arcpy_python, script_path],
            capture_output=True, text=True, timeout=300,
            encoding="utf-8", errors="replace", env=env,
        )
        return jsonify(
            ok=(result.returncode == 0),
            returncode=result.returncode,
            stdout=result.stdout or "",
            stderr=result.stderr or "",
            csvs=csv_tasks,
            parseErrors=parse_errors,
        )
    except subprocess.TimeoutExpired:
        return jsonify(friendly_error("arcpy_timeout")), 504
    except OSError as e:
        return jsonify(friendly_error("arcpy_exec_error", detail=str(e))), 500
    finally:
        if script_path and os.path.isfile(script_path):
            try:
                os.remove(script_path)
            except OSError:
                pass


@app.route("/api/smartsheet/batch-checkpoint", methods=["POST"])
def ss_batch_checkpoint():
    """Check if a checkpoint file exists for the given destination folder."""
    import json as _json
    body = request.json or {}
    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    ckpt_path = os.path.join(dest_folder, ".batch_checkpoint.json")
    if not os.path.isfile(ckpt_path):
        return jsonify(exists=False)
    try:
        with open(ckpt_path, "r", encoding="utf-8") as f:
            ckpt = _json.load(f)
        completed = len(ckpt.get("completed", {}))
        total = ckpt.get("total", 0)
        last_index = ckpt.get("last_index", -1)
        return jsonify(exists=True, completed=completed, total=total,
                       lastIndex=last_index)
    except Exception:
        return jsonify(exists=False)


def _scan_recent_shp_folders(base_dir: str, max_age_minutes: int = 90) -> list:
    """Scan base_dir for shapefile folders modified within the last max_age_minutes.
    Only walks into top-level subdirectories that are recent, avoiding full tree scan.
    Large batch downloads can take 30-60 min; 90 min covers that plus a safety margin.
    / Escanear base_dir buscando carpetas con shapefiles modificadas en los ultimos N minutos.
    Solo recorre subdirectorios de primer nivel recientes.
    """
    import time
    cutoff = time.time() - (max_age_minutes * 60)
    shp_folders = []
    if not os.path.isdir(base_dir):
        return shp_folders

    for entry in os.scandir(base_dir):
        if not entry.is_dir():
            continue
        # Check if this top-level subdirectory was recently modified
        # / Verificar si este subdirectorio de primer nivel fue modificado recientemente
        try:
            mtime = entry.stat().st_mtime
        except OSError:
            continue
        if mtime >= cutoff:
            # Walk recursively inside this recent subdirectory
            for dirpath, _dirnames, filenames in os.walk(entry.path):
                if any(f.lower().endswith(".shp") for f in filenames):
                    shp_folders.append(dirpath)
            continue

        # Level-2 check: scan child dirs even if parent mtime is old
        # (OneDrive sync may not update parent mtime)
        # / Nivel 2: revisar hijos aunque el padre no sea reciente
        try:
            for child in os.scandir(entry.path):
                if not child.is_dir():
                    continue
                try:
                    child_mtime = child.stat().st_mtime
                except OSError:
                    continue
                if child_mtime < cutoff:
                    continue
                for dirpath, _dirnames, filenames in os.walk(child.path):
                    if any(f.lower().endswith(".shp") for f in filenames):
                        shp_folders.append(dirpath)
        except OSError:
            continue

    return shp_folders


@app.route("/api/smartsheet/add-to-map-script", methods=["POST"])
def ss_add_to_map_script():
    """Generate ArcPy script to add downloaded shapefiles to ArcGIS Pro map.
    / Generar script ArcPy para agregar shapefiles descargados al mapa de ArcGIS Pro.

    Body: { destFolder, quarter, component, shpFolders? }
    If shpFolders is not provided, scans destFolder for subfolders containing .shp files.
    """
    from paso1_quarterly import script_add_shapefiles_to_map

    body = request.json or {}
    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    quarter = body.get("quarter", "").strip()
    comp = body.get("component", "C1")

    # Discover shapefile folders if not provided (only recent dirs)
    # / Descubrir carpetas con shapefiles si no se proporcionan (solo dirs recientes)
    shp_folders = body.get("shpFolders", [])
    if not shp_folders and os.path.isdir(dest_folder):
        shp_folders = _scan_recent_shp_folders(dest_folder)

    if not shp_folders:
        return jsonify(error="No shapefile folders found / No se encontraron carpetas con shapefiles"), 400

    aprx = os.getenv("APRX", "")
    map_name = os.getenv("AR_MAP_NAME", "AR_EbA_Area")

    if not aprx:
        return jsonify(error="APRX not configured in .env"), 400

    script = script_add_shapefiles_to_map({
        "aprx": aprx,
        "mapName": map_name,
        "shpFolders": shp_folders,
        "quarter": quarter,
        "component": comp,
        "lang": body.get("lang", "es"),
    })

    return jsonify(script=script, folders=len(shp_folders))


@app.route("/api/arcpy/close-pro", methods=["POST"])
def arcpy_close_pro():
    """Kill ArcGIS Pro process if running. / Cerrar ArcGIS Pro si está abierto."""
    import subprocess as _sp
    import time
    try:
        # Check if running / Verificar si está abierto
        check = _sp.run(
            ["tasklist", "/FI", "IMAGENAME eq ArcGISPro.exe", "/NH"],
            capture_output=True, text=True, timeout=10,
        )
        was_running = "ArcGISPro.exe" in check.stdout
        if was_running:
            _sp.run(["taskkill", "/F", "/IM", "ArcGISPro.exe"],
                     capture_output=True, timeout=10)
            # Wait until process is fully gone (up to 15s)
            # / Esperar hasta que el proceso haya terminado completamente (hasta 15s)
            for _ in range(6):
                time.sleep(2)
                recheck = _sp.run(
                    ["tasklist", "/FI", "IMAGENAME eq ArcGISPro.exe", "/NH"],
                    capture_output=True, text=True, timeout=10,
                )
                if "ArcGISPro.exe" not in recheck.stdout:
                    break
            # Extra pause for file lock release / Pausa extra para liberar bloqueos
            time.sleep(2)
        return jsonify(ok=True, wasRunning=was_running)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/api/arcpy/open-pro", methods=["POST"])
def arcpy_open_pro():
    """Reopen ArcGIS Pro with the project. / Reabrir ArcGIS Pro con el proyecto."""
    aprx = os.getenv("APRX", "")
    if not aprx or not os.path.isfile(aprx):
        return jsonify(ok=False, error="APRX not found"), 400
    try:
        os.startfile(aprx)
        return jsonify(ok=True, aprx=aprx)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500


@app.route("/api/arcpy/run-add-to-map", methods=["POST"])
def arcpy_run_add_to_map():
    """Generate and EXECUTE the ArcPy script that adds shapefiles to ArcGIS Pro.
    / Generar y EJECUTAR el script ArcPy que agrega shapefiles a ArcGIS Pro.

    Body: { destFolder, quarter, component, shpFolders? }
    Requires ARCPY_PYTHON to be set in .env.
    """
    import json as _json
    import tempfile
    from paso1_quarterly import script_add_shapefiles_to_map

    arcpy_python = os.getenv("ARCPY_PYTHON", "")
    if not arcpy_python or not os.path.isfile(arcpy_python):
        return jsonify(friendly_error("arcpy_not_configured")), 400

    body = request.json or {}
    try:
        dest_folder = str(safe_resolve(
            body.get("destFolder", str(ROOT / "downloads")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400

    quarter = body.get("quarter", "").strip()
    comp = body.get("component", "C1")

    # Discover shapefile folders (from batch or recent scan)
    # / Descubrir carpetas con shapefiles (del batch o escaneo reciente)
    seed_folders = body.get("shpFolders", [])
    if seed_folders:
        # Explicit folders from batch download: walk each seed
        # / Carpetas explícitas del batch download: recorrer cada semilla
        shp_folders = []
        for seed in seed_folders:
            if not os.path.isdir(seed):
                continue
            for dirpath, _dirnames, filenames in os.walk(seed):
                if any(f.lower().endswith(".shp") for f in filenames):
                    shp_folders.append(dirpath)
    else:
        # No explicit folders: scan only recently modified dirs (24h)
        # / Sin carpetas explícitas: escanear solo dirs modificados recientemente (24h)
        shp_folders = _scan_recent_shp_folders(dest_folder)

    if not shp_folders:
        return jsonify(error="No shapefile folders found / No se encontraron carpetas con shapefiles"), 400

    aprx = os.getenv("APRX", "")
    map_name = os.getenv("AR_MAP_NAME", "AR_EbA_Area")
    if not aprx:
        return jsonify(error="APRX not configured in .env"), 400

    script = script_add_shapefiles_to_map({
        "aprx": aprx,
        "mapName": map_name,
        "shpFolders": shp_folders,
        "quarter": quarter,
        "component": comp,
        "lang": body.get("lang", "es"),
    })

    # Write script to temp file and execute with ArcPy Python
    script_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".py", prefix="ar_add_to_map_",
            delete=False, encoding="utf-8",
        ) as tmp:
            tmp.write(script)
            script_path = tmp.name

        env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
        result = subprocess.run(
            [arcpy_python, script_path],
            capture_output=True, text=True, timeout=300,
            encoding="utf-8", errors="replace", env=env,
        )
        stdout = result.stdout or ""
        stderr = result.stderr or ""
        ok = result.returncode == 0

        return jsonify(
            ok=ok,
            returncode=result.returncode,
            stdout=stdout,
            stderr=stderr,
            script=script,
            folders=len(shp_folders),
        )
    except subprocess.TimeoutExpired:
        return jsonify(friendly_error("arcpy_timeout")), 504
    except OSError as e:
        return jsonify(friendly_error("arcpy_exec_error", detail=str(e))), 500
    finally:
        if script_path and os.path.isfile(script_path):
            try:
                os.remove(script_path)
            except OSError:
                pass



# ---------------------------------------------------------------------------
# API: Dafne M&E reception  (G11 + G7)
# ---------------------------------------------------------------------------

@app.route("/api/dafne/validate", methods=["POST"])
def dafne_validate():
    """
    POST /api/dafne/validate  body: {file_path}
    Validate Tbl_Integrado.xlsx required sheets.
    # ES: Valida las hojas requeridas en Tbl_Integrado.xlsx.
    """
    body = request.json or {}
    file_path = body.get("file_path", "").strip()
    if not file_path:
        return jsonify(error="file_path requerido / file_path required"), 400
    try:
        file_path = str(safe_resolve(file_path, _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    result = validate_integrado_xlsx(file_path)
    return jsonify(result)


@app.route("/api/dafne/place", methods=["POST"])
def dafne_place():
    """
    POST /api/dafne/place  body: {src_path, base_path, filename?}
    Copy Tbl_Integrado.xlsx to Power BI BasePath.
    # ES: Copia Tbl_Integrado.xlsx al BasePath de Power BI.
    """
    body = request.json or {}
    filename = body.get("filename", "Tbl_Integrado.xlsx").strip()
    try:
        src_path = str(safe_resolve(body.get("src_path", "").strip(), _allowed_roots()))
        base_path = str(safe_resolve(body.get("base_path", "").strip(), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    if not src_path or not base_path:
        return jsonify(error="src_path y base_path son requeridos / src_path and base_path required"), 400
    result = place_dafne_file(src_path, base_path, filename)
    return jsonify(result)


@app.route("/api/dafne/status")
def dafne_status():
    """
    GET /api/dafne/status?base_path=<path>
    Check if Tbl_Integrado.xlsx exists in BasePath.
    # ES: Verifica si Tbl_Integrado.xlsx existe en el BasePath.
    """
    base_path = request.args.get("base_path", "").strip()
    if not base_path:
        return jsonify(error="base_path requerido / base_path required"), 400
    try:
        base_path = str(safe_resolve(base_path, _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    result = get_integrado_status(base_path)
    return jsonify(result)


@app.route("/api/dafne/receive", methods=["POST"])
def dafne_receive():
    """
    POST /api/dafne/receive
    body: {file_path, quarter, base_path, metadata?}
    Integrated receive: validate + place + record history.
    # ES: Recibe, valida, coloca y registra el historial de recepción M&E.
    """
    body = request.json or {}
    file_path = body.get("file_path", "").strip()
    quarter = body.get("quarter", "").strip()
    base_path = body.get("base_path", "").strip()
    metadata = body.get("metadata") or {}
    if not file_path or not quarter or not base_path:
        return jsonify(
            error="file_path, quarter y base_path son requeridos / file_path, quarter and base_path required"
        ), 400
    result = receive_me_data(file_path, quarter, base_path, metadata)
    # ES: Si la recepción fue exitosa, marcar substep 5b como completado
    # EN: On successful reception, mark substep 5b as complete
    if result["success"]:
        _pipeline_state.set_substep("5b", "success")
    return jsonify(result)


@app.route("/api/dafne/history")
def dafne_history():
    """
    GET /api/dafne/history?quarter=<quarter>
    Return reception history for the given quarter, newest first.
    # ES: Devuelve el historial de recepciones del trimestre.
    """
    quarter = request.args.get("quarter", "").strip()
    if not quarter:
        return jsonify(error="quarter requerido / quarter required"), 400
    history = get_reception_history(quarter)
    return jsonify(history)


@app.route("/api/dafne/quarters")
def dafne_quarters():
    """
    GET /api/dafne/quarters
    List all quarters with reception history.
    # ES: Lista los trimestres con historial de recepción.
    """
    quarters = list_reception_quarters()
    return jsonify(quarters)


@app.route("/api/orchestrator/advance-substep", methods=["POST"])
def orch_advance_substep():
    """
    POST /api/orchestrator/advance-substep  body: {key, status}
    Manually advance a substep (e.g. {key:"5a", status:"success"}).
    # ES: Avanza manualmente un subestado del orquestador.
    """
    body = request.json or {}
    key = body.get("key", "").strip()
    status = body.get("status", "success").strip()
    if not key:
        return jsonify(error="key requerido / key required"), 400
    try:
        _pipeline_state.set_substep(key, status)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    return jsonify(_pipeline_state.to_dict())


# ---------------------------------------------------------------------------
# API: Excel Master  (PASO 5)
# ---------------------------------------------------------------------------

@app.route("/api/paso5/generate-excel", methods=["POST"])
def paso5_generate_excel():
    """Generate unified Excel master from Smartsheet + GIS data."""
    try:
        import openpyxl
    except ImportError:
        return jsonify(error="openpyxl not installed. Run: pip install openpyxl"), 500

    body = request.json or {}
    dest_folder = body.get("destFolder", str(ROOT / "downloads"))
    os.makedirs(dest_folder, exist_ok=True)

    wb = openpyxl.Workbook()

    for comp in ["C1", "C2", "C3"]:
        sid = _sheet_id(comp)
        if not sid:
            continue
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
        if not r.ok:
            continue
        sheet = r.json()
        col_map = {c["id"]: c["title"] for c in sheet.get("columns", [])}
        col_titles = [c["title"] for c in sheet.get("columns", [])]

        ws = wb.create_sheet(title=comp)
        ws.append(col_titles)
        for row in sheet.get("rows", []):
            cell_map = {c.get("columnId"): c.get("value") for c in row.get("cells", [])}
            values = [cell_map.get(col["id"], "") for col in sheet.get("columns", [])]
            ws.append(values)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"AR_Master_{ts}.xlsx"
    filepath = os.path.join(dest_folder, filename)
    wb.save(filepath)

    return jsonify(file=filename, path=filepath, sheets=wb.sheetnames)


# ---------------------------------------------------------------------------
# API: Publish-to-Web link check  (G10)
# ---------------------------------------------------------------------------

@app.route("/api/pbi/publish-check")
def pbi_publish_check():
    """
    GET /api/pbi/publish-check?url=<embed_url>
    Check whether a Power BI Publish-to-Web embed URL is publicly accessible.
    # ES: Verifica si un enlace "Publicar en Web" de Power BI es accesible.
    # EN: Returns {"status", "http_code", "message"}.
    """
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify(error="Parámetro 'url' requerido / 'url' parameter required"), 400
    result = check_publish_url(url)
    return jsonify(result)


@app.route("/api/pbi/embed-url")
def pbi_embed_url():
    """
    GET /api/pbi/embed-url?report_id=<id>&group_id=<id>
    Retrieve embed URL via Power BI REST API (requires Azure AD credentials).
    # ES: No disponible — requiere credenciales Azure AD (ver paso6_powerbi.py).
    # EN: Not available — requires Azure AD credentials (see paso6_powerbi.py).
    """
    report_id = request.args.get("report_id", "").strip()
    group_id = request.args.get("group_id", "").strip() or None
    if not report_id:
        return jsonify(error="Parámetro 'report_id' requerido / 'report_id' parameter required"), 400
    try:
        result = get_embed_url_via_api(report_id=report_id, group_id=group_id)
        return jsonify(result)
    except NotImplementedError as exc:
        return jsonify(
            error="API no disponible / API not available",
            detail=str(exc),
        ), 501


# ---------------------------------------------------------------------------
# API: PBI PDF Export  (PASO 7 — G9)
# ---------------------------------------------------------------------------

@app.route("/api/pbi/export-pdf", methods=["POST"])
def pbi_export_pdf():
    """
    POST /api/pbi/export-pdf
    Generate a Power BI PDF export guide (Strategy B — no Azure AD credentials).
    / Genera una guía de exportación PDF para Power BI (Estrategia B — sin credenciales Azure AD).

    Body: {
        "report_name": str,   // e.g. "AR_Dashboard"
        "output_path": str,   // suggested save folder
        "quarter":     str,   // e.g. "T2026_Q2"
        "include_mcp": bool   // include MCP DAX verification steps (default true)
    }

    Returns: {
        "guide": str,          // bilingual step-by-step guide text
        "strategy": "manual",
        "api_available": false
    }
    """
    from paso6_powerbi import generate_pdf_export_guide

    body = request.json or {}
    guide = generate_pdf_export_guide(body)

    # EN: Record the export event in pipeline state for data summary.
    # ES: Registrar el evento de exportación en el estado del pipeline para el resumen.
    _pipeline_state.record_pdf_export(
        strategy="manual",
        pdf_path=body.get("output_path", ""),
        quarter=body.get("quarter", ""),
    )

    return jsonify(
        guide=guide,
        strategy="manual",
        api_available=False,
        note=(
            "Power BI Service REST API requiere credenciales Azure AD no configuradas. "
            "/ Power BI Service REST API requires Azure AD credentials not configured."
        ),
    )


# ---------------------------------------------------------------------------
# API: Reports  (PASO 7)
# ---------------------------------------------------------------------------

@app.route("/api/paso7/error-report")
def paso7_error_report():
    report = generate_error_report(
        _pipeline_state.to_dict(),
    )
    return jsonify(report)


@app.route("/api/paso7/data-summary")
def paso7_data_summary():
    summary = generate_data_summary(
        _pipeline_state.to_dict(),
    )
    return jsonify(summary)


@app.route("/api/paso7/backup", methods=["POST"])
def paso7_backup():
    body = request.json or {}
    if not body.get("sourceFolder", ""):
        return jsonify(error="sourceFolder required"), 400
    try:
        source = str(safe_resolve(body["sourceFolder"], _allowed_roots()))
        backup = str(safe_resolve(
            body.get("backupFolder", str(ROOT / "downloads" / "backups")), _allowed_roots()))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    result = manage_backups(source, backup)
    return jsonify(result)


# ---------------------------------------------------------------------------
# API: Orchestrator
# ---------------------------------------------------------------------------

@app.route("/api/orchestrator/start", methods=["POST"])
def orch_start():
    body = request.json or {}
    pasos = body.get("pasos")
    result = _orchestrator.start_pipeline(pasos)
    return jsonify(result)


@app.route("/api/orchestrator/status")
def orch_status():
    return jsonify(_orchestrator.get_status())


@app.route("/api/orchestrator/advance", methods=["POST"])
def orch_advance():
    body = request.json or {}
    paso = body.get("paso")
    if not paso:
        return jsonify(error="paso required"), 400
    result = _orchestrator.advance_paso(int(paso))
    return jsonify(result)


@app.route("/api/orchestrator/retry", methods=["POST"])
def orch_retry():
    body = request.json or {}
    paso = body.get("paso")
    if not paso:
        return jsonify(error="paso required"), 400
    result = _orchestrator.retry_paso(int(paso))
    return jsonify(result)


@app.route("/api/smartsheet/export-csv", methods=["POST"])
def ss_export_csv():
    """Export Smartsheet data as CSV for ArcGIS import."""
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    dest_folder = body.get("destFolder", str(ROOT / "downloads"))
    sid = _sheet_id(comp)
    if not sid:
        return jsonify(friendly_error("no_sheet_id")), 400

    os.makedirs(dest_folder, exist_ok=True)

    try:
        r = smartsheet_request("GET", f"{SS_BASE}/sheets/{sid}", headers=_ss_headers(), timeout=30)
    except requests.ConnectionError:
        return jsonify(friendly_error("ss_network")), 503
    if r.status_code == 401:
        return jsonify(friendly_error("ss_401")), 401
    if not r.ok:
        return jsonify(friendly_error("unknown")), r.status_code
    sheet = r.json()
    col_map = {c["id"]: c["title"] for c in sheet.get("columns", [])}
    col_titles = [c["title"] for c in sheet.get("columns", [])]

    ts = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"ss_{comp}_{ts}.csv"
    filepath = os.path.join(dest_folder, filename)

    def cell_val_by_id(row, cid):
        for c in row.get("cells", []):
            if c.get("columnId") == cid:
                return c.get("value")
        return None

    col_map_title_to_id = {c["title"]: c["id"] for c in sheet.get("columns", [])}
    target_rows = [
        row for row in sheet.get("rows", [])
        if row_start <= row.get("rowNumber", 0) <= row_end
    ]
    if comp == "C2":
        target_rows = _c2_filter_rows(target_rows, body, col_map_title_to_id, cell_val_by_id)

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(col_titles)
        for row in target_rows:
            values = []
            cell_map = {c.get("columnId"): c.get("value") for c in row.get("cells", [])}
            for col in sheet.get("columns", []):
                values.append(cell_map.get(col["id"], ""))
            writer.writerow(values)

    return jsonify(file=filename, path=filepath, rows=row_end - row_start + 1)


# ---------------------------------------------------------------------------
# API: Cleaning Criteria (PASO 3 — G4)
# EN: Store and retrieve coordinator-decided cleaning criteria per quarter.
# ES: Almacena y recupera los criterios de limpieza decididos por el coordinador.
# ---------------------------------------------------------------------------

@app.route("/api/criteria/")
def criteria_list_quarters():
    """
    GET /api/criteria/
    EN: List all quarters that have saved criteria.
    ES: Lista todos los trimestres con criterios guardados.
    """
    return jsonify(quarters=list_criteria_quarters())


@app.route("/api/criteria/<quarter>", methods=["GET", "POST"])
def criteria_quarter(quarter: str):
    """
    GET  /api/criteria/<quarter>  — load criteria for a quarter
    POST /api/criteria/<quarter>  — overwrite all criteria entries
    ES: GET carga; POST sobreescribe todos los criterios del trimestre.
    """
    if request.method == "POST":
        body = request.get_json(silent=True) or {}
        entries = body.get("entries", [])
        if not isinstance(entries, list):
            return jsonify(friendly_error("unknown", **{})), 400
        data = save_criteria(quarter, entries)
        return jsonify(data)
    return jsonify(load_criteria(quarter))


@app.route("/api/criteria/<quarter>/add", methods=["POST"])
def criteria_add_entry(quarter: str):
    """
    POST /api/criteria/<quarter>/add
    EN: Append a single criterion entry to the quarter's list.
    ES: Agrega un único criterio a la lista del trimestre.
    Body: {layer, reason, decision, cdg_actividad, recorded_by?}
    """
    body = request.get_json(silent=True) or {}
    required = ["layer", "reason", "decision"]
    missing = [k for k in required if not body.get(k, "").strip()]
    if missing:
        return jsonify({
            "error": True,
            "message": f"Campos requeridos faltantes / Required fields missing: {', '.join(missing)}",
        }), 400
    entry = add_criterion(quarter, body)
    return jsonify(entry), 201


@app.route("/api/criteria/<quarter>/export-pdf")
def criteria_export_pdf(quarter: str):
    """
    GET /api/criteria/<quarter>/export-pdf
    EN: Generate and download a PDF (or HTML fallback) of the criteria list.
    ES: Genera y descarga un PDF (o HTML de respaldo) con la lista de criterios.
    """
    import tempfile

    suffix = ".pdf"
    with tempfile.NamedTemporaryFile(
        suffix=suffix, prefix=f"criteria_{quarter}_", delete=False
    ) as tmp:
        tmp_path = tmp.name

    generated = export_criteria_pdf(quarter, tmp_path)

    ext = pathlib.Path(generated).suffix.lower()
    mime = "application/pdf" if ext == ".pdf" else "text/html"
    download_name = f"Criterios_{quarter}{ext}"

    from flask import send_file
    return send_file(generated, mimetype=mime, as_attachment=True, download_name=download_name)


# ---------------------------------------------------------------------------
# Cleaning Stats — PASO 3 before/after comparison (G5)
# ES: Almacena y recupera estadísticas de limpieza (antes/después del Erase).
# EN: Store and retrieve cleaning stats (before/after Erase pipeline).
# ---------------------------------------------------------------------------

@app.route("/api/pipeline/cleaning-stats", methods=["POST"])
def cleaning_stats_save():
    """
    POST /api/pipeline/cleaning-stats
    Body: {
      "before": {"count": int, "area_ha": float},
      "after":  {"count": int, "area_ha": float},
      "removed": [{"cdg": str, "area_ha": float, "reason": str}, ...]
    }
    """
    body = request.json or {}
    before = body.get("before", {})
    after = body.get("after", {})
    removed = body.get("removed", [])

    stats = {
        "before": {
            "count": int(before.get("count", 0)),
            "area_ha": float(before.get("area_ha", 0.0)),
        },
        "after": {
            "count": int(after.get("count", 0)),
            "area_ha": float(after.get("area_ha", 0.0)),
        },
        "removed": removed,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
    }
    _pipeline_state.set_cleaning_stats(stats)
    return jsonify(ok=True, stats=stats)


@app.route("/api/pipeline/cleaning-stats", methods=["GET"])
def cleaning_stats_get():
    """GET /api/pipeline/cleaning-stats — return stored cleaning stats."""
    return jsonify(_pipeline_state.cleaning_stats)


# ---------------------------------------------------------------------------
# API: M&E Comparison Engine  (PASO 6 / G8)
# ---------------------------------------------------------------------------

@app.route("/api/compare/me", methods=["POST"])
def compare_me():
    """
    POST /api/compare/me
    Body: {integrado_path, pbi_source, quarter, metrics?}
    Loads Dafne values from integrado_path, loads PBI values from pbi_source,
    and runs compare_me_values().

    # ES: Ejecuta la comparación entre valores Dafne y Power BI para el trimestre.
    # EN: Runs the Dafne vs Power BI comparison for the given quarter.
    """
    body = request.json or {}
    integrado_path = body.get("integrado_path", "")
    pbi_source = body.get("pbi_source", {"type": "inline", "output": {}, "outcome": {}, "impact": {}})
    quarter = body.get("quarter", "")
    metrics = body.get("metrics") or None

    dafne_data = load_dafne_values(integrado_path)
    if dafne_data.get("error"):
        return jsonify(error=dafne_data["error"]), 400

    pbi_data = load_pbi_values(pbi_source)
    if pbi_data.get("error"):
        return jsonify(error=pbi_data["error"]), 400

    comparison = compare_me_values(dafne_data, pbi_data, metrics)
    comparison["quarter"] = quarter
    return jsonify(comparison)


@app.route("/api/compare/me/drill", methods=["POST"])
def compare_me_drill():
    """
    POST /api/compare/me/drill
    Body: {cdg_actividad, integrado_path, pbi_source, quarter?}
    Returns field-level discrepancy for a single CdgActvdd.

    # ES: Profundiza en la discrepancia a nivel de campo para una actividad.
    # EN: Drills into field-level discrepancy for one activity.
    """
    body = request.json or {}
    cdg = body.get("cdg_actividad", "")
    integrado_path = body.get("integrado_path", "")
    pbi_source = body.get("pbi_source", {"type": "inline", "output": {}, "outcome": {}, "impact": {}})

    dafne_data = load_dafne_values(integrado_path)
    if dafne_data.get("error"):
        return jsonify(error=dafne_data["error"]), 400

    pbi_data = load_pbi_values(pbi_source)
    if pbi_data.get("error"):
        return jsonify(error=pbi_data["error"]), 400

    dafne_row = dafne_data.get("rows", {}).get(cdg, {})
    pbi_row = pbi_data.get("rows", {}).get(cdg, {})
    result = find_discrepancy_cause(dafne_row, pbi_row, cdg)
    return jsonify(result)


@app.route("/api/compare/me/decision", methods=["POST"])
def compare_me_decision():
    """
    POST /api/compare/me/decision
    Body: {metric, dafne_val, pbi_val, decision, manual_val?, quarter}
    Persists the final value decision.

    # ES: Guarda la decisión del valor final (Dafne/PBI/manual) para una métrica.
    # EN: Saves the final value decision (Dafne/PBI/manual) for a metric.
    """
    body = request.json or {}
    metric = body.get("metric", "")
    dafne_val = body.get("dafne_val", 0)
    pbi_val = body.get("pbi_val", 0)
    decision = body.get("decision", "dafne")
    manual_val = body.get("manual_val")
    quarter = body.get("quarter", "")

    try:
        entry = decide_final_value(
            metric=metric,
            dafne_val=float(dafne_val),
            pbi_val=float(pbi_val),
            decision=decision,
            manual_val=float(manual_val) if manual_val is not None else None,
            quarter=quarter,
        )
    except (ValueError, TypeError) as exc:
        return jsonify(error=str(exc)), 400

    return jsonify(ok=True, entry=entry)


@app.route("/api/compare/me/report")
def compare_me_report():
    """
    GET /api/compare/me/report?quarter=...
    Returns stored comparison decisions + summary for the quarter.

    # ES: Devuelve el reporte de decisiones de comparación para un trimestre.
    # EN: Returns the comparison decisions report for a quarter.
    """
    quarter = request.args.get("quarter", "")
    report = get_comparison_report(quarter)
    return jsonify(report)


# ---------------------------------------------------------------------------
# API: AGOL Connection — OAuth 2.0 PKCE Authentication
# EN: Operator authenticates via AGOL OAuth popup — no passwords on server.
# ES: El operador se autentica via OAuth AGOL — sin contraseñas en el servidor.
# ---------------------------------------------------------------------------

@app.route("/api/agol/auth/start", methods=["POST"])
def agol_auth_start():
    """Start OAuth 2.0 PKCE authentication flow.

    Returns {url, state} — the frontend opens url in a popup.
    / Inicia el flujo de autenticación OAuth 2.0 PKCE.
    """
    import agol_connect as ac

    body = request.get_json(silent=True) or {}
    org_url = (body.get("org_url") or "").strip() or None

    redirect_uri = f"http://localhost:{request.host.split(':')[-1]}{ac.REDIRECT_PATH}"
    try:
        result = ac.build_authorize_url(
            redirect_uri=redirect_uri,
            org_url=org_url,
        )
        return jsonify(result)
    except ValueError as exc:
        app.logger.error("[AGOL] OAuth start error: %s", exc)
        return jsonify(friendly_error("p0_agol_error",
                       detail=str(exc))), 400


@app.route("/oauth/callback")
def agol_oauth_callback():
    """OAuth 2.0 callback — exchanges code for token, closes popup.

    / Callback OAuth 2.0 — intercambia código por token y cierra el popup.
    """
    import agol_connect as ac

    code = request.args.get("code", "")
    state = request.args.get("state", "")
    error = request.args.get("error", "")

    if error:
        app.logger.warning("[AGOL] OAuth error returned: %s", error)
        return (
            "<html><body><h3>AGOL OAuth Error</h3>"
            f"<p>{error}</p>"
            "<script>window.close();</script></body></html>"
        ), 400

    if not code or not state:
        return (
            "<html><body><h3>Missing code or state</h3>"
            "<script>window.close();</script></body></html>"
        ), 400

    redirect_uri = f"http://localhost:{request.host.split(':')[-1]}{ac.REDIRECT_PATH}"
    try:
        info = ac.exchange_code(code=code, state=state, redirect_uri=redirect_uri)
        app.logger.info("[AGOL] OAuth success: %s @ %s",
                        info.get("username"), info.get("org"))
        # Render a page that notifies the opener and closes itself
        return (
            "<html><body>"
            "<h3>Conectado a AGOL / Connected to AGOL</h3>"
            "<p>Esta ventana se cerrará automáticamente.</p>"
            "<script>"
            "if (window.opener) { window.opener.postMessage({agolConnected: true}, '*'); }"
            "setTimeout(function(){ window.close(); }, 1500);"
            "</script></body></html>"
        )
    except Exception as exc:
        app.logger.error("[AGOL] OAuth token exchange failed: %s", exc)
        return (
            "<html><body><h3>Error de autenticación</h3>"
            f"<p>{type(exc).__name__}</p>"
            "<script>window.close();</script></body></html>"
        ), 500


@app.route("/api/agol/status", methods=["GET"])
def agol_status():
    """Check current AGOL connection status.

    / Verificar estado de conexión AGOL actual.
    """
    import agol_connect as ac
    info = ac.get_connection_info()
    info["env_org_url"] = os.getenv("ARCGIS_ORG_URL", "")
    info["env_has_client_id"] = bool(os.getenv("ARCGIS_CLIENT_ID", ""))
    return jsonify(info)


@app.route("/api/agol/disconnect", methods=["POST"])
def agol_disconnect():
    """Clear AGOL session and OAuth tokens from memory.

    / Limpiar sesión AGOL y tokens OAuth de memoria.
    """
    import agol_connect as ac
    ac.clear_caches()
    app.logger.info("[AGOL] Disconnected — OAuth tokens cleared from memory")
    return jsonify({"ok": True, "connected": False})


# ---------------------------------------------------------------------------
# API: PASO 0 — AGOL vs Smartsheet Diagnostic / Diagnóstico AGOL vs Smartsheet
# EN: Compares ArcGIS Online Feature Layer (previous-quarter verified data)
#     against Smartsheet (all data including new entries) to categorize
#     activities as: verified-match, verified-mismatch, or new (incomplete).
# ES: Compara Feature Layer de ArcGIS Online (datos verificados del trimestre
#     anterior) contra Smartsheet (todos los datos) para categorizar
#     actividades como: verificado-coincide, verificado-discrepancia, o nuevo.
# ---------------------------------------------------------------------------

# ── AGOL URL SSRF defense ───────────────────────────────────────────────
_ALLOWED_AGOL_HOSTS = frozenset({
    "services.arcgis.com",
    "services1.arcgis.com", "services2.arcgis.com",
    "services3.arcgis.com", "services4.arcgis.com",
    "services5.arcgis.com", "services6.arcgis.com",
    "services7.arcgis.com", "services8.arcgis.com",
    "services9.arcgis.com",
})


def _is_allowed_agol_url(url: str) -> bool:
    """Validate URL points to a known ArcGIS Online domain."""
    try:
        parsed = urllib.parse.urlparse(url)
    except Exception:
        return False
    if parsed.scheme not in ("https", "http"):
        return False
    host = (parsed.hostname or "").lower()
    if host in _ALLOWED_AGOL_HOSTS:
        return True
    if host.endswith(".arcgis.com"):
        return True
    return False


# ── AGOL response cache ──────────────────────────────────────────────────
_AGOL_CACHE: dict = {}        # url -> {"ts": float, "data": dict}
_AGOL_CACHE_TTL = 300         # 5 minutes

@app.route("/api/paso0/diagnose", methods=["POST"])
def paso0_diagnose():
    """
    Body JSON:
      - component:  "C1" | "C2" | "C3"
      - rowStart, rowEnd: optional row range
      - c2GrantType, c2Contrato: optional C2 filters
      - agol_polygon_url: override polygon Feature Layer URL
      - agol_point_url:   override point Feature Layer URL
      - ha_threshold: optional tolerance for hectares comparison (default 5.0)

    Two AGOL layers (polygon + point) are fetched and merged by CdgActvdd.
    Area_ha values from both layers are summed per activity code.

    Returns:
      { summary, comparison, groups, diagnostics }
    """
    body = request.json or {}
    comp = body.get("component", "C1")
    row_start = body.get("rowStart", 1)
    row_end = body.get("rowEnd", 9999)
    ha_threshold = float(body.get("ha_threshold", 5.0))
    cfg = _cfg()

    # ── Resolve AGOL layer URLs ──────────────────────────────────────
    # Priority: dashboard URL override > authenticated Item ID > .env URL
    # / Prioridad: URL del dashboard > Item ID autenticado > URL de .env
    import agol_connect as ac

    _use_authenticated = False
    _agol_fl_polygon = None   # FeatureLayer object (if authenticated)
    _agol_fl_point = None

    agol_polygon_url = body.get("agol_polygon_url", "")
    agol_point_url = body.get("agol_point_url", "")

    if not agol_polygon_url and not agol_point_url and ac.is_connected():
        # Try to resolve via authenticated Item IDs
        polygon_item_id = os.getenv("AGOL_POLYGON_ITEM_ID", "")
        point_item_id = os.getenv("AGOL_POINT_ITEM_ID", "")
        if polygon_item_id or point_item_id:
            try:
                if polygon_item_id:
                    _agol_fl_polygon, agol_polygon_url = ac.resolve_layer(polygon_item_id)
                if point_item_id:
                    _agol_fl_point, agol_point_url = ac.resolve_layer(point_item_id)
                _use_authenticated = True
                app.logger.info("[P0] Using authenticated AGOL session (Item IDs)")
            except (LookupError, IndexError, Exception) as exc:
                app.logger.warning("[P0] Item ID resolution failed, falling back to URLs: %s", exc)

    # Fallback to .env URLs if not resolved above
    if not agol_polygon_url:
        agol_polygon_url = cfg.get("AGOL_POLYGON_URL", "")
    if not agol_point_url:
        agol_point_url = cfg.get("AGOL_POINT_URL", "")

    app.logger.info("[P0] diagnose START  comp=%s  rows=%d–%d  threshold=%.2f",
                    comp, row_start, row_end, ha_threshold)

    sid = _sheet_id(comp)
    if not sid:
        app.logger.warning("[P0] No sheet ID configured for %s", comp)
        return jsonify(friendly_error("no_sheet_id")), 400

    if not agol_polygon_url and not agol_point_url:
        app.logger.warning("[P0] No AGOL Feature Layer URLs configured")
        return jsonify(friendly_error("p0_no_agol_url")), 400

    # SSRF defense: reject non-AGOL domains (skip for authenticated Item ID mode
    # because URLs are resolved server-side from trusted AGOL metadata)
    if not _use_authenticated:
        for _label, _url in [("polygon", agol_polygon_url), ("point", agol_point_url)]:
            if _url and not _is_allowed_agol_url(_url):
                return jsonify(error=f"AGOL {_label} URL must be an ArcGIS Online domain (*.arcgis.com)"), 400

    # ── 1. Fetch Smartsheet data ─────────────────────────────────────────
    cache_key = f"{comp}:{sid}"
    cached = _SS_CACHE.get(cache_key)
    ss_cache_hit = cached and (time.time() - cached["ts"]) < _SS_CACHE_TTL
    if ss_cache_hit:
        ss_payload = cached["data"]
        app.logger.info("[P0] Smartsheet cache HIT for %s (age %.0fs)", comp, time.time() - cached["ts"])
    else:
        app.logger.info("[P0] Fetching Smartsheet sheet %s (id=%s) …", comp, sid)
        try:
            r = smartsheet_request(
                "GET", f"{SS_BASE}/sheets/{sid}",
                headers=_ss_headers(),
                params={"include": "attachments,objectValue"},
                timeout=30,
            )
        except requests.ConnectionError:
            app.logger.error("[P0] Smartsheet connection error for %s", comp)
            return jsonify(friendly_error("ss_network")), 503
        if r.status_code == 401:
            app.logger.error("[P0] Smartsheet 401 — token inválido")
            return jsonify(friendly_error("ss_401")), 401
        if not r.ok:
            app.logger.error("[P0] Smartsheet HTTP %d: %s", r.status_code, r.text[:200])
            return jsonify(friendly_error("unknown")), r.status_code
        data = r.json()
        col_map = {c["title"]: c["id"] for c in data.get("columns", [])}
        rows_out = []
        for row in data.get("rows", []):
            cells = {}
            for cell in row.get("cells", []):
                for title, cid in col_map.items():
                    if cell.get("columnId") == cid:
                        v = cell.get("value")
                        if v is None:
                            obj = cell.get("objectValue")
                            if isinstance(obj, dict):
                                v = obj.get("value") or obj.get("number")
                            elif obj is not None:
                                v = obj
                        if v is None:
                            v = cell.get("displayValue")
                        cells[title] = v
            rows_out.append({
                "rowNumber": row.get("rowNumber"),
                "id": row.get("id"),
                "cells": cells,
                "attachments": [
                    {"id": a["id"], "name": a.get("name", "")}
                    for a in row.get("attachments", [])
                ],
            })
        ss_payload = {
            "name": data.get("name"),
            "totalRows": len(rows_out),
            "columns": list(col_map.keys()),
            "rows": rows_out,
        }
        _SS_CACHE[cache_key] = {"ts": time.time(), "data": ss_payload}
        app.logger.info("[P0] Sheet '%s' fetched: %d rows, %d columns",
                        ss_payload.get("name", "?"), len(rows_out), len(col_map))

    all_rows = ss_payload.get("rows", [])
    columns = ss_payload.get("columns", [])
    app.logger.info("[P0] Sheet columns: %s", columns)

    # ── 2. Filter rows by range / visible row numbers and C2 filters ────
    row_numbers = body.get("rowNumbers")
    if row_numbers and isinstance(row_numbers, list):
        visible_set = set(row_numbers)
        rows = [r for r in all_rows if r.get("rowNumber", 0) in visible_set]
        app.logger.info("[P0] Filtering by %d visible row numbers from Data View", len(visible_set))
    else:
        rows = [r for r in all_rows if row_start <= r.get("rowNumber", 0) <= row_end]
    if comp == "C2":
        grant_filter = body.get("c2GrantType", "")
        contrato_filter = body.get("c2Contrato", "")
        if grant_filter or contrato_filter:
            # EN: NÚMERO DE CONTRATO is set only on parent/group-header rows.
            #     Child activity rows inherit the contrato from the nearest
            #     preceding parent row (same pattern used in step-5 grouping).
            # ES: NÚMERO DE CONTRATO solo aparece en filas padre/encabezado.
            #     Las filas hijas heredan el contrato de la fila padre más
            #     cercana anterior (mismo patrón del agrupamiento en paso 5).
            filtered = []
            current_contrato = ""
            for r in rows:
                contrato = r["cells"].get("NÚMERO DE CONTRATO", "") or ""
                if contrato:
                    current_contrato = contrato
                if contrato_filter:
                    if current_contrato != contrato_filter:
                        continue
                elif grant_filter:
                    if grant_filter not in str(current_contrato):
                        continue
                filtered.append(r)
            rows = filtered

    # ── 3. Pre-extract SS codes for targeted AGOL query ────────────────
    # EN: Extract activity codes from filtered SS rows BEFORE querying AGOL.
    #     This allows us to build targeted CdgActvdd IN (...) queries
    #     instead of fetching all AGOL data for the component.
    # ES: Extraer códigos de actividad de las filas SS filtradas ANTES de
    #     consultar AGOL, para construir consultas dirigidas.
    HA_COL = "TOTAL DE HECTÁREAS"
    CODE_COL = "CÓDIGO DE LA ACTIVIDAD"

    ss_target_codes = set()
    for r in rows:
        cells = r.get("cells", {})
        code = cells.get(CODE_COL)
        fecha = cells.get("FECHA DE LA ACTIVIDAD")
        if not fecha and not code:
            continue
        code = str(code or "").strip()
        ss_ha_raw = cells.get(HA_COL)
        if ss_ha_raw is not None:
            try:
                float(str(ss_ha_raw).replace(",", "."))
            except (ValueError, TypeError):
                continue
            if code:
                ss_target_codes.add(code)

    app.logger.info("[P0] SS target codes extracted: %d unique codes from %d filtered rows",
                    len(ss_target_codes), len(rows))

    # ── 4. Fetch AGOL Feature Layers (targeted by SS codes) ─────────────
    # EN: Query AGOL only for codes found in the filtered SS rows.
    #     If authenticated via agol_connect, uses FeatureLayer.query() with
    #     automatic token management. Otherwise falls back to requests.post().
    # ES: Consultar AGOL solo para los códigos encontrados en las filas SS.
    #     Si está autenticado, usa FeatureLayer.query() con gestión automática
    #     de token. Si no, usa requests.post() como antes.
    AGOL_BATCH_SIZE = 100  # codes per AGOL query batch

    agol_layers = []  # [(name, url, feature_layer_or_None), ...]
    if agol_polygon_url:
        agol_layers.append(("polygon", agol_polygon_url, _agol_fl_polygon))
    if agol_point_url:
        agol_layers.append(("point", agol_point_url, _agol_fl_point))

    agol_records = {}       # {code: area_ha}  — merged from all layers
    agol_cache_hit = True   # False if any layer was fetched fresh
    agol_layer_stats = {}   # {layer_name: feature_count}

    def _resolve_query_url(layer_url):
        """Resolve AGOL layer query URL from FeatureServer base."""
        import agol_connect as _ac
        token = _ac.get_token()

        query_url = layer_url.rstrip("/")
        if query_url.endswith("/FeatureServer"):
            try:
                params = {"f": "json"}
                if token:
                    params["token"] = token
                fs_meta = requests.get(query_url, params=params, timeout=15).json()
                layers = fs_meta.get("layers", [])
                lid = layers[0]["id"] if layers else 0
            except Exception:
                lid = 0
            query_url += f"/{lid}/query"
        elif not query_url.endswith("/query"):
            query_url += "/query"
        return query_url

    def _agol_fetch_paginated(query_url, where_clause, out_fields, layer_name):
        """Fetch all features from AGOL with pagination support.

        If an OAuth token is available it is included so that private
        layers can be queried even without FeatureLayer objects.
        """
        import agol_connect as _ac
        token = _ac.get_token()

        post_data = {
            "where": where_clause,
            "outFields": out_fields,
            "returnGeometry": "false",
            "f": "json",
        }
        if token:
            post_data["token"] = token

        all_features = []
        agol_r = requests.post(query_url, data=post_data, timeout=90)
        if not agol_r.ok:
            app.logger.warning("[P0] AGOL %s HTTP %d: %s",
                               layer_name, agol_r.status_code, agol_r.text[:200])
            return None
        try:
            resp_json = agol_r.json()
        except Exception:
            app.logger.warning("[P0] AGOL %s empty/invalid JSON response", layer_name)
            return None
        if "error" in resp_json:
            err = resp_json["error"]
            app.logger.warning("[P0] AGOL %s error: %s", layer_name, err)
            # Token Required → signal caller so it can abort early
            if isinstance(err, dict) and err.get("code") in (498, 499):
                return "TOKEN_REQUIRED"
            return None
        features = resp_json.get("features", [])
        all_features.extend(features)
        exceeded = resp_json.get("exceededTransferLimit", False)
        page = 1
        while exceeded:
            page += 1
            offset = len(all_features)
            app.logger.info("[P0] AGOL %s page %d (offset=%d) …",
                            layer_name, page, offset)
            try:
                next_data = {
                    "where": where_clause,
                    "outFields": out_fields,
                    "returnGeometry": "false",
                    "f": "json",
                    "resultOffset": str(offset),
                }
                if token:
                    next_data["token"] = token
                next_r = requests.post(query_url, data=next_data, timeout=90)
                if not next_r.ok:
                    break
                try:
                    next_json = next_r.json()
                except Exception:
                    break
                if "error" in next_json:
                    break
                batch = next_json.get("features", [])
                if not batch:
                    break
                all_features.extend(batch)
                exceeded = next_json.get("exceededTransferLimit", False)
            except (requests.ConnectionError, requests.Timeout):
                app.logger.warning("[P0] AGOL %s pagination stopped at page %d",
                                   layer_name, page)
                break
        if page > 1:
            app.logger.info("[P0] AGOL %s: %d pages, %d total features",
                            layer_name, page, len(all_features))
        return all_features

    def _agol_fetch_authenticated(fl, where_clause, out_fields, layer_name):
        """Fetch features using authenticated FeatureLayer (SDK handles token/pagination)."""
        try:
            features = ac.query_features(fl, where=where_clause, out_fields=out_fields)
            app.logger.info("[P0] AGOL %s (authenticated): %d features for where=%s",
                            layer_name, len(features), where_clause[:80])
            return features
        except Exception as exc:
            app.logger.warning("[P0] AGOL %s authenticated query failed: %s", layer_name, exc)
            return None

    def _agol_fetch(layer_name, fl, query_url, where_clause, out_fields):
        """Route to authenticated or anonymous fetch."""
        if fl is not None:
            return _agol_fetch_authenticated(fl, where_clause, out_fields, layer_name)
        return _agol_fetch_paginated(query_url, where_clause, out_fields, layer_name)

    def _features_to_records(features):
        """Convert AGOL features to {code: area_ha} dict, grouped by CdgActvdd."""
        records = {}
        for feat in features:
            attrs = feat.get("attributes", {})
            code = str(attrs.get("CdgActvdd") or attrs.get("cdgactvdd") or "").strip()
            area = 0.0
            raw_area = attrs.get("Area_ha") or attrs.get("area_ha")
            if raw_area is not None:
                try:
                    area = float(raw_area)
                except (ValueError, TypeError):
                    pass
            if code:
                records[code] = records.get(code, 0.0) + area
        return records

    # Build targeted IN-clause batches from SS codes
    code_list = sorted(ss_target_codes)
    code_batches = [code_list[i:i + AGOL_BATCH_SIZE]
                    for i in range(0, len(code_list), AGOL_BATCH_SIZE)] if code_list else []

    for layer_name, layer_url, fl in agol_layers:
        # ── Cache check (keyed by layer URL + sorted codes hash) ────────
        codes_key = "|".join(code_list) if len(code_list) <= 500 else f"n={len(code_list)}"
        cache_key_targeted = f"{layer_url}|IN:{hash(codes_key)}"
        cached_layer = _AGOL_CACHE.get(cache_key_targeted)
        if cached_layer and (time.time() - cached_layer["ts"]) < _AGOL_CACHE_TTL:
            layer_records = cached_layer["data"]
            app.logger.info("[P0] AGOL %s cache HIT (age %.0fs), %d codes (targeted query)",
                            layer_name, time.time() - cached_layer["ts"],
                            len(layer_records))
        else:
            agol_cache_hit = False
            app.logger.info("[P0] Querying AGOL %s (targeted: %d codes in %d batches, auth=%s): %s",
                            layer_name, len(code_list), len(code_batches),
                            fl is not None, layer_url)
            layer_records = {}
            query_url = _resolve_query_url(layer_url) if fl is None else None
            try:
                if code_batches:
                    # Targeted query: CdgActvdd IN (...) per batch
                    token_required = False
                    for bi, batch in enumerate(code_batches):
                        in_values = ",".join(f"'{c}'" for c in batch)
                        where_clause = f"CdgActvdd IN ({in_values})"
                        app.logger.info("[P0] AGOL %s batch %d/%d (%d codes)",
                                        layer_name, bi + 1, len(code_batches), len(batch))
                        features = _agol_fetch(layer_name, fl, query_url,
                                               where_clause, "CdgActvdd,Area_ha")
                        if features == "TOKEN_REQUIRED":
                            token_required = True
                            break
                        if features is None:
                            app.logger.warning("[P0] AGOL %s batch %d failed, trying broad fallback",
                                               layer_name, bi + 1)
                            features = []
                        batch_records = _features_to_records(features)
                        for code, area in batch_records.items():
                            layer_records[code] = layer_records.get(code, 0.0) + area
                    if token_required:
                        return jsonify(friendly_error(
                            "p0_agol_error",
                            detail="Token Required – haga clic en 'Conectar AGOL' "
                                   "para autenticarse antes de ejecutar el diagnóstico."
                        )), 502
                else:
                    # No SS codes → broad fallback query
                    comp_num = comp.replace("C", "")
                    agol_output_value = f"output{comp_num}"
                    where_clause = f"output='{agol_output_value}'"
                    app.logger.info("[P0] AGOL %s no SS codes, broad query: %s",
                                    layer_name, where_clause)
                    features = _agol_fetch(layer_name, fl, query_url,
                        where_clause, "CdgActvdd,Area_ha,output")
                    if features == "TOKEN_REQUIRED":
                        return jsonify(friendly_error(
                            "p0_agol_error",
                            detail="Token Required – haga clic en 'Conectar AGOL' "
                                   "para autenticarse antes de ejecutar el diagnóstico."
                        )), 502
                    if features is None:
                        # Last resort: 1=1
                        features = _agol_fetch(layer_name, fl, query_url,
                            "1=1", "CdgActvdd,Area_ha")
                    if features == "TOKEN_REQUIRED" or features is None:
                        app.logger.error("[P0] AGOL %s: all query attempts failed", layer_name)
                        return jsonify(friendly_error("p0_agol_error",
                                       detail=f"{layer_name}: All query attempts failed")), 502
                    layer_records = _features_to_records(features)

                app.logger.info("[P0] AGOL %s: %d codes, %.2f ha (targeted=%s, auth=%s)",
                                layer_name, len(layer_records),
                                sum(layer_records.values()),
                                bool(code_batches), fl is not None)
            except requests.ConnectionError as exc:
                app.logger.error("[P0] AGOL %s connection error: %s", layer_name, exc)
                return jsonify(friendly_error("p0_agol_error",
                               detail=f"{layer_name}: No se pudo conectar")), 502
            except requests.Timeout:
                app.logger.error("[P0] AGOL %s timeout", layer_name)
                return jsonify(friendly_error("p0_agol_error",
                               detail=f"{layer_name}: Timeout")), 504
            _AGOL_CACHE[cache_key_targeted] = {"ts": time.time(), "data": layer_records}

        agol_layer_stats[layer_name] = len(layer_records)
        # Merge into combined dict (groupby CdgActvdd, sum Area_ha)
        for code, area in layer_records.items():
            agol_records[code] = agol_records.get(code, 0.0) + area

    app.logger.info("[P0] AGOL merged: %d unique codes, %.2f ha total (layers: %s)",
                    len(agol_records), sum(agol_records.values()), agol_layer_stats)

    # ── Debug: show sample AGOL codes for comparison diagnosis ──────────
    _agol_sample = sorted(agol_records.keys())[:8]
    if _agol_sample:
        app.logger.info("[P0] AGOL sample codes: %s", _agol_sample)

    # ── 5. Build Smartsheet row diagnostics ──────────────────────────────
    # EN: Only rows with TOTAL DE HECTÁREAS are comparison targets.
    # ES: Solo las filas con TOTAL DE HECTÁREAS son objetivo de comparación.

    row_diagnostics = []
    skipped_empty = 0
    skipped_no_ha = 0
    agol_codes_set = set(agol_records.keys())

    for r in rows:
        cells = r.get("cells", {})
        code = cells.get(CODE_COL)
        fecha = cells.get("FECHA DE LA ACTIVIDAD")

        # Skip header/empty rows (no date and no code)
        if not fecha and not code:
            skipped_empty += 1
            continue

        code = str(code or "").strip()

        # Parse SS hectares
        ss_ha_raw = cells.get(HA_COL)
        ss_ha = None
        if ss_ha_raw is not None:
            try:
                ss_ha = float(str(ss_ha_raw).replace(",", "."))
            except (ValueError, TypeError):
                ss_ha = None

        # Skip rows without hectares — not comparison targets
        if ss_ha is None:
            skipped_no_ha += 1
            continue

        # Determine comparison status
        agol_ha = agol_records.get(code)
        if agol_ha is not None:
            # Code exists in AGOL (previous quarter verified)
            diff = abs(ss_ha - agol_ha)
            if diff <= ha_threshold:
                cmp_status = "verified_match"
            else:
                cmp_status = "verified_mismatch"
        else:
            # Code NOT in AGOL → new this quarter, needs review
            cmp_status = "new"

        row_diagnostics.append({
            "rowNumber": r.get("rowNumber"),
            "rowId": r.get("id"),
            "code": code,
            "date": str(fecha or ""),
            "ss_ha": round(ss_ha, 4) if ss_ha is not None else None,
            "agol_ha": round(agol_ha, 4) if agol_ha is not None else None,
            "ha_diff": round(abs(ss_ha - agol_ha), 4) if agol_ha is not None else None,
            "cmp_status": cmp_status,
            # C2 grouping fields
            "contrato": cells.get("NÚMERO DE CONTRATO", "") or "",
            "org": cells.get("ORGANIZACIÓN", "") or "",
            "quarter": cells.get("TRIMESTRE QUE REPORTA", "") or "",
        })

    # Codes in AGOL but NOT in Smartsheet (filtered set)
    ss_codes_set = {rd["code"] for rd in row_diagnostics if rd["code"]}
    agol_only = sorted(agol_codes_set - ss_codes_set)

    # ── Debug: show sample SS codes for comparison diagnosis ────────────
    _ss_sample = sorted(ss_codes_set)[:8]
    if _ss_sample:
        app.logger.info("[P0] SS sample codes: %s", _ss_sample)

    # Log analysis stats
    cnt_match = sum(1 for rd in row_diagnostics if rd["cmp_status"] == "verified_match")
    cnt_mismatch = sum(1 for rd in row_diagnostics if rd["cmp_status"] == "verified_mismatch")
    cnt_new = sum(1 for rd in row_diagnostics if rd["cmp_status"] == "new")
    app.logger.info("[P0] SS rows: %d in range, %d skipped empty, %d skipped no-ha, %d analyzed",
                    len(rows), skipped_empty, skipped_no_ha, len(row_diagnostics))
    app.logger.info("[P0] Comparison: match=%d, mismatch=%d, new=%d, agol_only=%d",
                    cnt_match, cnt_mismatch, cnt_new, len(agol_only))
    if cnt_new == len(row_diagnostics) and len(row_diagnostics) > 0 and len(agol_records) > 0:
        app.logger.warning(
            "[P0] 0%% match: none of the %d SS codes matched %d AGOL codes. "
            "Check code format. SS[0]=%r  AGOL[0]=%r",
            len(ss_codes_set), len(agol_records),
            _ss_sample[0] if _ss_sample else "?",
            _agol_sample[0] if _agol_sample else "?",
        )

    # ── 6. Group rows (for C2: by Contrato–Quarter) ────────────────────
    groups = []
    if comp == "C2":
        from collections import OrderedDict
        from ar_utils import is_shape_zip_attachment as _is_shp_zip
        group_map = OrderedDict()
        last_contrato = ""
        last_org = ""
        last_quarter = ""
        for rd in row_diagnostics:
            if rd["contrato"]:
                last_contrato = rd["contrato"]
            if rd["org"]:
                last_org = rd["org"]
            if rd["quarter"]:
                last_quarter = rd["quarter"]
            grant = "PPD" if "PPD" in last_contrato else ("PMD" if "PMD" in last_contrato else "")
            key = f"{last_contrato}|{last_quarter}"
            if key not in group_map:
                group_map[key] = {
                    "key": key,
                    "grant_type": grant,
                    "contrato": last_contrato,
                    "org": last_org,
                    "quarter": last_quarter,
                    "rows": [],
                    "summaryRowNumber": None,
                }
            group_map[key]["rows"].append(rd)

        # Find summary rows (with SHP attachments) for each C2 group.
        # Summary rows have CONTRATO + TRIMESTRE set and shape-zip attachments.
        # They are typically skipped by row_diagnostics (no code/date).
        # Fallback: if no summary row has SHP, check child rows with
        # TOTAL DE HECTÁREAS that have SHP attachments.
        # / Buscar filas resumen (con adjuntos SHP) para cada grupo C2.
        #   Fallback: si la fila resumen no tiene SHP, buscar filas hijas con
        #   hectáreas que tengan SHP adjuntos.
        HA_COL_LOCAL = "TOTAL DE HECTÁREAS"
        for r in all_rows:
            cells = r.get("cells", {})
            contrato = cells.get("NÚMERO DE CONTRATO", "") or ""
            qt = cells.get("TRIMESTRE QUE REPORTA", "") or ""
            if not contrato or not qt:
                continue
            has_shp = any(
                _is_shp_zip(a.get("name", ""))
                for a in r.get("attachments", [])
            )
            if not has_shp:
                continue
            key = f"{contrato}|{qt}"
            if key in group_map and not group_map[key]["summaryRowNumber"]:
                group_map[key]["summaryRowNumber"] = r.get("rowNumber")

        # Fallback: scan child rows (with hectares + SHP) for groups still
        # missing a summaryRowNumber.
        _groups_missing = {k for k, g in group_map.items() if not g["summaryRowNumber"]}
        if _groups_missing:
            last_contrato = ""
            last_qt = ""
            for r in all_rows:
                cells = r.get("cells", {})
                c = cells.get("NÚMERO DE CONTRATO", "") or ""
                q = cells.get("TRIMESTRE QUE REPORTA", "") or ""
                if c:
                    last_contrato = c
                if q:
                    last_qt = q
                key = f"{last_contrato}|{last_qt}"
                if key not in _groups_missing:
                    continue
                has_shp = any(
                    _is_shp_zip(a.get("name", ""))
                    for a in r.get("attachments", [])
                )
                if not has_shp:
                    continue
                group_map[key]["summaryRowNumber"] = r.get("rowNumber")
                _groups_missing.discard(key)

        groups = list(group_map.values())
    else:
        groups = [{
            "key": comp,
            "grant_type": "",
            "contrato": "",
            "org": "",
            "quarter": "",
            "rows": row_diagnostics,
        }]

    # ── 7. Group-level summaries ─────────────────────────────────────────
    for g in groups:
        g_rows = g["rows"]
        g["total"] = len(g_rows)
        g["verified_match"] = sum(1 for rd in g_rows if rd["cmp_status"] == "verified_match")
        g["verified_mismatch"] = sum(1 for rd in g_rows if rd["cmp_status"] == "verified_mismatch")
        g["new"] = sum(1 for rd in g_rows if rd["cmp_status"] == "new")
        g["verified_pct"] = round(g["verified_match"] / max(g["total"], 1) * 100)
        g["ss_ha_total"] = round(sum(rd["ss_ha"] for rd in g_rows if rd["ss_ha"] is not None), 2)
        g["agol_ha_total"] = round(sum(rd["agol_ha"] for rd in g_rows if rd["agol_ha"] is not None), 2)

    # ── 8. Build global summary ──────────────────────────────────────────
    total = len(row_diagnostics)
    verified_pct = round(cnt_match / max(total, 1) * 100)
    ss_ha_sum = round(sum(rd["ss_ha"] for rd in row_diagnostics if rd["ss_ha"] is not None), 2)
    agol_ha_sum = round(sum(rd["agol_ha"] for rd in row_diagnostics if rd["agol_ha"] is not None), 2)

    comparison = {
        "agol_total_codes": len(agol_records),
        "agol_total_ha": round(sum(agol_records.values()), 2),
        "ss_total_codes": len(ss_codes_set),
        "ss_total_ha": ss_ha_sum,
        "verified_match": cnt_match,
        "verified_mismatch": cnt_mismatch,
        "new_this_quarter": cnt_new,
        "agol_only": agol_only,
        "agol_only_count": len(agol_only),
        "verified_pct": verified_pct,
        "ha_threshold": ha_threshold,
    }

    # ── 9. Build diagnostics detail for frontend log ─────────────────────
    diagnostics = {
        "sheet_name": ss_payload.get("name", ""),
        "sheet_total_rows": len(all_rows),
        "rows_in_range": len(rows),
        "skipped_empty": skipped_empty,
        "skipped_no_ha": skipped_no_ha,
        "analyzed": total,
        "agol_polygon_url": agol_polygon_url,
        "agol_point_url": agol_point_url,
        "agol_layer_stats": agol_layer_stats,
        "agol_features": len(agol_records),
        "ss_cache_hit": ss_cache_hit,
        "agol_cache_hit": agol_cache_hit,
        "agol_targeted": bool(code_batches),
        "agol_authenticated": _use_authenticated,
        "agol_query_batches": len(code_batches),
        "ss_target_codes_count": len(ss_target_codes),
        "ss_sample_codes": sorted(ss_codes_set)[:5],
        "agol_sample_codes": sorted(agol_records.keys())[:5],
    }

    result = {
        "component": comp,
        "sheet_name": ss_payload.get("name", ""),
        "summary": {
            "total_rows": total,
            "verified_match": cnt_match,
            "verified_mismatch": cnt_mismatch,
            "new_this_quarter": cnt_new,
            "verified_pct": verified_pct,
            "ss_ha_total": ss_ha_sum,
            "agol_ha_total": agol_ha_sum,
        },
        "comparison": comparison,
        "groups": groups,
        "diagnostics": diagnostics,
    }
    app.logger.info("[P0] diagnose END  comp=%s  verified=%d%%  match=%d  mismatch=%d  new=%d",
                    comp, verified_pct, cnt_match, cnt_mismatch, cnt_new)
    return jsonify(result)


@app.route("/api/paso0/csv", methods=["POST"])
def paso0_csv():
    """Generate a CSV download with the full Paso 0 comparison detail.

    Accepts the same body as /api/paso0/diagnose but returns text/csv.
    Columns: Row#, CÓDIGO DE LA ACTIVIDAD, Fecha, SS_ha, AGOL_ha, Diff_ha, Estado,
             Contrato, Organización, Trimestre
    A second section lists AGOL-only codes not found in the Smartsheet range.
    """
    import io
    import csv
    from flask import Response

    # Re-use diagnose logic (call internal)
    with app.test_request_context(
        "/api/paso0/diagnose",
        method="POST",
        json=request.json or {},
        content_type="application/json",
    ):
        resp = paso0_diagnose()
        # resp is (response, status_code) tuple on error, or Response on success
        if isinstance(resp, tuple):
            return resp
        data = resp.get_json()

    if not data or "groups" not in data:
        return jsonify(friendly_error("unknown")), 500

    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    comp = data.get("component", "")
    sheet = data.get("sheet_name", "")
    cmp_info = data.get("comparison", {})
    summary = data.get("summary", {})

    writer.writerow(["# Diagnóstico AGOL vs Smartsheet / AGOL vs Smartsheet Diagnostic"])
    writer.writerow(["# Componente / Component", comp])
    writer.writerow(["# Hoja / Sheet", sheet])
    writer.writerow(["# Tolerancia ha / Threshold (ha)", cmp_info.get("ha_threshold", 5.0)])
    writer.writerow(["# Verificado-coincide / Verified-match", summary.get("verified_match", 0)])
    writer.writerow(["# Discrepancia / Mismatch", summary.get("verified_mismatch", 0)])
    writer.writerow(["# Nuevos / New", summary.get("new_this_quarter", 0)])
    writer.writerow(["# AGOL-only", cmp_info.get("agol_only_count", 0)])
    writer.writerow([])

    # Detail rows
    writer.writerow([
        "Fila / Row#",
        "CÓDIGO DE LA ACTIVIDAD",
        "Fecha / Date",
        "SS_ha",
        "AGOL_ha",
        "Diff_ha",
        "Estado / Status",
        "Contrato",
        "Organización / Organization",
        "Trimestre / Quarter",
    ])

    status_labels = {
        "verified_match": "Verificado-Coincide / Verified-Match",
        "verified_mismatch": "Discrepancia / Mismatch",
        "new": "Nuevo / New",
    }

    for g in data.get("groups", []):
        for r in g.get("rows", []):
            writer.writerow([
                r.get("rowNumber", ""),
                r.get("code", ""),
                r.get("date", ""),
                r.get("ss_ha") if r.get("ss_ha") is not None else "",
                r.get("agol_ha") if r.get("agol_ha") is not None else "",
                r.get("ha_diff") if r.get("ha_diff") is not None else "",
                status_labels.get(r.get("cmp_status", ""), r.get("cmp_status", "")),
                r.get("contrato", ""),
                r.get("org", ""),
                r.get("quarter", ""),
            ])

    # AGOL-only section
    agol_only = cmp_info.get("agol_only", [])
    if agol_only:
        writer.writerow([])
        writer.writerow(["# Códigos solo en AGOL (no en Smartsheet filtrado) / Codes only in AGOL"])
        writer.writerow(["CÓDIGO / Code", "AGOL_ha"])
        # agol_only is just a list of codes; individual ha not available in summary
        for code in agol_only:
            writer.writerow([code, ""])

    csv_bytes = buf.getvalue()
    filename = f"paso0_{comp}_{sheet}.csv".replace(" ", "_")
    return Response(
        csv_bytes,
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ---------------------------------------------------------------------------
# Paso 1 Agent Loop endpoints
# ---------------------------------------------------------------------------

@app.route("/api/paso1-agent/state", methods=["GET"])
def paso1_agent_get_state():
    dest = request.args.get("destFolder", "")
    if not dest:
        return jsonify(friendly_error("agent_no_dest")), 400
    try:
        state = paso1_agent.load_agent_state(dest, _allowed_roots())
    except ValueError:
        return jsonify(friendly_error("agent_update_fail", detail="Invalid destFolder")), 400
    if state is None:
        return jsonify({"initialized": False})
    return jsonify({"initialized": True, **state})


@app.route("/api/paso1-agent/init", methods=["POST"])
def paso1_agent_init():
    data = request.get_json(force=True)
    component = data.get("component", "C1")
    quarter = data.get("quarter", "")
    dest = data.get("destFolder", "")
    diagnose_data = data.get("diagnoseData")
    if not dest or not diagnose_data:
        return jsonify(friendly_error("agent_missing_params")), 400
    try:
        state = paso1_agent.init_agent_state(
            dest, component, quarter, diagnose_data, _allowed_roots(),
        )
    except ValueError as e:
        return jsonify(friendly_error("agent_update_fail", detail=str(e))), 400
    return jsonify(state)


@app.route("/api/paso1-agent/mark-complete", methods=["POST"])
def paso1_agent_mark_complete():
    data = request.get_json(force=True)
    dest = data.get("destFolder", "")
    item_id = data.get("itemId", "")
    step = data.get("stepReached", "done")
    try:
        state = paso1_agent.update_item(
            dest, item_id,
            {"item_state": "done", "step_reached": step},
            _allowed_roots(),
        )
    except (ValueError, FileNotFoundError, KeyError) as e:
        return jsonify(friendly_error("agent_update_fail", detail=str(e))), 400
    return jsonify(state)


@app.route("/api/paso1-agent/mark-skip", methods=["POST"])
def paso1_agent_mark_skip():
    data = request.get_json(force=True)
    dest = data.get("destFolder", "")
    item_id = data.get("itemId", "")
    reason = data.get("reason", "user_skipped")
    try:
        state = paso1_agent.update_item(
            dest, item_id,
            {"item_state": "skipped", "step_reached": reason},
            _allowed_roots(),
        )
    except (ValueError, FileNotFoundError, KeyError) as e:
        return jsonify(friendly_error("agent_update_fail", detail=str(e))), 400
    return jsonify(state)


@app.route("/api/paso1-agent/mark-pending", methods=["POST"])
def paso1_agent_mark_pending():
    data = request.get_json(force=True)
    dest = data.get("destFolder", "")
    item_id = data.get("itemId", "")
    try:
        state = paso1_agent.update_item(
            dest, item_id,
            {"item_state": "pending", "step_reached": None},
            _allowed_roots(),
        )
    except (ValueError, FileNotFoundError, KeyError) as e:
        return jsonify(friendly_error("agent_update_fail", detail=str(e))), 400
    return jsonify(state)


@app.route("/api/paso1-agent/reset", methods=["POST"])
def paso1_agent_reset():
    data = request.get_json(force=True)
    dest = data.get("destFolder", "")
    try:
        paso1_agent.reset_agent_state(dest, _allowed_roots())
    except ValueError as e:
        return jsonify(friendly_error("agent_update_fail", detail=str(e))), 400
    return jsonify({"reset": True})


# ---------------------------------------------------------------------------
# Global error handlers (S3)
# ---------------------------------------------------------------------------

@app.errorhandler(500)
def handle_500(e):
    # EN: Internal server error — catch-all for unhandled exceptions
    return jsonify(friendly_error("unknown")), 500


@app.errorhandler(404)
def handle_404(e):
    # EN: Route not found
    return jsonify(friendly_error("unknown")), 404


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Local token: {_LOCAL_TOKEN}")
    print("AR Workflow Server starting on http://localhost:5000")
    print("Open this URL in your browser to use the dashboard.")
    app.run(host="127.0.0.1", port=5000, debug=os.getenv("FLASK_DEBUG", "0") == "1",
            threaded=True)
